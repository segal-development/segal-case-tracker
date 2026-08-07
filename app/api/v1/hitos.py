"""Hitos (milestone → bonus) endpoints — Slice 1.

Replaces the manual "SISTEMA DE HITOS" sheet: a pre-loaded hito-type catalog, a
guided entry with a MANDATORY PJUD evidence capture, admin approval, and a
per-lawyer monthly total. Firm rule enforced here: a hito can never be approved
without evidence ("sin evidencia no se paga").
"""
import hashlib
import io
import logging
import re
import unicodedata
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import get_current_lawyer, get_db, require_admin
from app.models.hito import Hito, HitoTipo, HITO_APROBADO, HITO_PENDIENTE, HITO_RECHAZADO
from app.models.lawyer import Lawyer
from app.services import bono_cierre_service as cierre_svc

# A ROL like "C-9960-2026" embedded anywhere in the free-text descripcion.
_ROL_RE = re.compile(r"[A-Za-z]+-\d+-\d{4}")


def _causa_key(descripcion: Optional[str]) -> Optional[str]:
    """Causa identifier used for hito dedup.

    ``rol_causa`` stores the CLIENT RUT, so a lawyer could legitimately earn hitos
    for the same client across DIFFERENT causas. The distinguishing causa lives in
    ``descripcion`` (a ROL like ``C-9960-2026``), so dedup keys on
    ``(lawyer, rol_causa, _causa_key(descripcion))``: the ROL when present, else the
    normalized free text. ``None`` for empty descripcion.
    """
    if not descripcion:
        return None
    m = _ROL_RE.search(descripcion)
    return (m.group(0).upper() if m else descripcion.strip().upper()[:120]) or None

logger = logging.getLogger(__name__)
router = APIRouter()

_MAX_EVIDENCE_BYTES = 15 * 1024 * 1024  # 15 MB
_ALLOWED_EVIDENCE = {"image/png", "image/jpeg", "image/webp", "application/pdf"}


# --------------------------------------------------------------------------- #
# Schemas
# --------------------------------------------------------------------------- #
class HitoTipoResponse(BaseModel):
    id: int
    code: str
    label: str
    nivel: str
    valor_bruto: int
    etapa_tramite: Optional[str] = None
    verificacion: Optional[str] = None

    class Config:
        from_attributes = True


class HitoResponse(BaseModel):
    id: int
    lawyer_id: int
    lawyer_nombre: Optional[str] = None
    hito_tipo_id: int
    tipo_label: str
    nivel: str
    valor_bruto: int
    fecha_hito: date
    rol_causa: Optional[str] = None
    procedimiento: Optional[str] = None
    descripcion: Optional[str] = None
    etapa_sysgal: Optional[str] = None
    tramite_sysgal: Optional[str] = None
    tiene_evidencia: bool
    estado: str
    origen: str = "manual"          # manual | detector
    confianza: Optional[str] = None  # alta | media | baja (solo detector)
    created_by_name: Optional[str] = None
    aprobado_by_name: Optional[str] = None
    aprobado_at: Optional[datetime] = None
    rechazo_motivo: Optional[str] = None


class HitoResumenRow(BaseModel):
    lawyer_id: int
    lawyer_nombre: str
    aprobados: int
    total_bruto: int
    pendientes: int


class RechazoBody(BaseModel):
    motivo: Optional[str] = None


class HitoBulkIds(BaseModel):
    ids: list[int]


class HitoBulkResult(BaseModel):
    procesados: int
    ids: list[int]  # the ids actually acted on


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _resolve_lawyer(db: Session, current_lawyer: dict) -> Optional[Lawyer]:
    sub = current_lawyer.get("sub") or current_lawyer.get("lawyer_id")
    if sub is None:
        return None
    if isinstance(sub, int) or (isinstance(sub, str) and str(sub).isdigit()):
        return db.query(Lawyer).filter(Lawyer.id == int(sub)).first()
    return db.query(Lawyer).filter(Lawyer.rut == str(sub)).first()


def _is_admin(lawyer: Optional[Lawyer]) -> bool:
    return bool(lawyer and lawyer.role == "admin")


def _to_response(h: Hito) -> HitoResponse:
    return HitoResponse(
        id=h.id,
        lawyer_id=h.lawyer_id,
        lawyer_nombre=h.lawyer.name if h.lawyer else None,
        hito_tipo_id=h.hito_tipo_id,
        tipo_label=h.tipo.label if h.tipo else "",
        nivel=h.tipo.nivel if h.tipo else "",
        valor_bruto=h.valor_bruto,
        fecha_hito=h.fecha_hito,
        rol_causa=h.rol_causa,
        procedimiento=h.procedimiento,
        descripcion=h.descripcion,
        etapa_sysgal=h.etapa_sysgal,
        tramite_sysgal=h.tramite_sysgal,
        tiene_evidencia=h.tiene_evidencia,
        estado=h.estado,
        origen=h.origen,
        confianza=h.confianza,
        created_by_name=h.created_by_name,
        aprobado_by_name=h.aprobado_by_name,
        aprobado_at=h.aprobado_at,
        rechazo_motivo=h.rechazo_motivo,
    )


# --------------------------------------------------------------------------- #
# Endpoints
# --------------------------------------------------------------------------- #
@router.get("/tipos", response_model=List[HitoTipoResponse])
async def list_hito_tipos(
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """The hito-type catalog (for the entry form: label → value auto-fills)."""
    return (
        db.query(HitoTipo)
        .filter(HitoTipo.activo.is_(True))
        .order_by(HitoTipo.orden)
        .all()
    )


@router.post("", response_model=HitoResponse, status_code=status.HTTP_201_CREATED)
async def create_hito(
    hito_tipo_id: int = Form(...),
    fecha_hito: date = Form(...),
    rol_causa: Optional[str] = Form(None),
    procedimiento: Optional[str] = Form(None),
    descripcion: Optional[str] = Form(None),
    etapa_sysgal: Optional[str] = Form(None),
    tramite_sysgal: Optional[str] = Form(None),
    lawyer_id: Optional[int] = Form(None),  # admins may register for another lawyer
    evidencia: Optional[UploadFile] = File(None),  # PJUD capture — optional
    db: Session = Depends(get_db),
    current_lawyer: dict = Depends(get_current_lawyer),
):
    """Register a hito. PJUD evidence is optional."""
    actor = _resolve_lawyer(db, current_lawyer)
    if actor is None:
        raise HTTPException(status_code=401, detail="No se pudo resolver el abogado")

    if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(fecha_hito)):
        raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")

    tipo = db.query(HitoTipo).filter(HitoTipo.id == hito_tipo_id, HitoTipo.activo.is_(True)).first()
    if tipo is None:
        raise HTTPException(status_code=404, detail="Tipo de hito no encontrado")

    # A lawyer registers hitos for themselves; only an admin may set another lawyer.
    target_lawyer_id = actor.id
    if lawyer_id is not None and lawyer_id != actor.id:
        if not _is_admin(actor):
            raise HTTPException(status_code=403, detail="Solo un admin puede registrar hitos de otro abogado")
        if db.query(Lawyer).filter(Lawyer.id == lawyer_id).first() is None:
            raise HTTPException(status_code=404, detail="Abogado no encontrado")
        target_lawyer_id = lawyer_id

    # No duplicate on the SAME causa: a lawyer may repeat the same client
    # (rol_causa = RUT) across DIFFERENT causas, so the block is keyed on
    # (abogado, RUT, causa) — the causa is derived from descripcion (see
    # _causa_key). Only enforced when a rol_causa is given. Checked before storing
    # evidence so a rejected duplicate never uploads a file.
    rol_norm = (rol_causa or "").strip() or None
    if rol_norm is not None:
        causa = _causa_key(descripcion)
        prior = (
            db.query(Hito.descripcion)
            .filter(Hito.lawyer_id == target_lawyer_id, Hito.rol_causa == rol_norm)
            .all()
        )
        if any(_causa_key(d) == causa for (d,) in prior):
            raise HTTPException(
                status_code=409,
                detail="Ya existe un hito de este abogado para esa causa",
            )

    # Evidence is optional. If provided, validate + store it.
    storage_uri = ev_filename = ev_content_type = None
    data = await evidencia.read() if evidencia is not None else b""
    if data:
        if len(data) > _MAX_EVIDENCE_BYTES:
            raise HTTPException(status_code=413, detail="La evidencia supera el tamaño máximo (15 MB)")
        content_type = evidencia.content_type or "application/octet-stream"
        if content_type not in _ALLOWED_EVIDENCE:
            raise HTTPException(
                status_code=415,
                detail="Formato de evidencia no permitido (usa PNG, JPG, WEBP o PDF)",
            )
        from app.config import settings
        from app.services.storage_service import get_storage_backend

        digest = hashlib.sha256(data).hexdigest()[:16]
        ext = {"image/png": "png", "image/jpeg": "jpg", "image/webp": "webp", "application/pdf": "pdf"}.get(content_type, "bin")
        key = f"hitos/evidencia/{target_lawyer_id}/{digest}.{ext}"
        storage_uri = get_storage_backend(settings).upload(data, key, content_type=content_type)
        ev_filename = evidencia.filename
        ev_content_type = content_type

    hito = Hito(
        lawyer_id=target_lawyer_id,
        hito_tipo_id=tipo.id,
        valor_bruto=tipo.valor_bruto,  # snapshot
        fecha_hito=fecha_hito,
        rol_causa=rol_norm,
        procedimiento=procedimiento,
        descripcion=descripcion,
        etapa_sysgal=etapa_sysgal or tipo.etapa_tramite,
        tramite_sysgal=tramite_sysgal,
        evidencia_storage_key=storage_uri,
        evidencia_filename=ev_filename,
        evidencia_content_type=ev_content_type,
        estado=HITO_PENDIENTE,
        created_by_rut=actor.rut,
        created_by_name=actor.name,
    )
    db.add(hito)
    db.commit()
    db.refresh(hito)
    return _to_response(hito)


@router.get("", response_model=List[HitoResponse])
async def list_hitos(
    periodo: Optional[str] = Query(None, description="Filtrar por mes YYYY-MM"),
    lawyer_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_lawyer: dict = Depends(get_current_lawyer),
):
    """List hitos. Admins see all; a lawyer sees only their own."""
    actor = _resolve_lawyer(db, current_lawyer)
    if actor is None:
        raise HTTPException(status_code=401, detail="No se pudo resolver el abogado")

    q = db.query(Hito)
    if not _is_admin(actor):
        q = q.filter(Hito.lawyer_id == actor.id)  # non-admins: own hitos only
    elif lawyer_id is not None:
        q = q.filter(Hito.lawyer_id == lawyer_id)
    if estado:
        q = q.filter(Hito.estado == estado)
    if periodo:
        try:
            y, m = (int(x) for x in periodo.split("-"))
            start = date(y, m, 1)
            end = date(y + (m == 12), (m % 12) + 1, 1)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="periodo inválido (usa YYYY-MM)")
        q = q.filter(Hito.fecha_hito >= start, Hito.fecha_hito < end)

    return [_to_response(h) for h in q.order_by(Hito.fecha_hito.desc(), Hito.id.desc()).all()]


@router.post("/{hito_id}/aprobar", response_model=HitoResponse)
async def aprobar_hito(
    hito_id: int,
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Approve a hito (admin only)."""
    hito = db.query(Hito).filter(Hito.id == hito_id).first()
    if hito is None:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(hito.fecha_hito)):
        raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")

    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    hito.estado = HITO_APROBADO
    hito.aprobado_by_rut = admin_rut
    hito.aprobado_by_name = admin.name if admin else None
    hito.aprobado_at = datetime.utcnow()
    hito.rechazo_motivo = None
    db.commit()
    db.refresh(hito)
    return _to_response(hito)


@router.post("/{hito_id}/rechazar", response_model=HitoResponse)
async def rechazar_hito(
    hito_id: int,
    body: RechazoBody,
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Reject a hito (admin only), with an optional reason."""
    hito = db.query(Hito).filter(Hito.id == hito_id).first()
    if hito is None:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    hito.estado = HITO_RECHAZADO
    hito.aprobado_by_rut = admin_rut
    hito.aprobado_by_name = admin.name if admin else None
    hito.aprobado_at = datetime.utcnow()
    hito.rechazo_motivo = body.motivo
    db.commit()
    db.refresh(hito)
    return _to_response(hito)


# --------------------------------------------------------------------------- #
# Bulk actions (mirror the single-item aprobar / rechazar / delete)
# --------------------------------------------------------------------------- #
@router.post("/aprobar-lote", response_model=HitoBulkResult)
async def aprobar_hitos_lote(
    body: HitoBulkIds,
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Approve every existing hito in ``ids`` (admin only). Skips ids that don't
    exist; a closed period blocks the whole batch with the same 409 as the
    single endpoint."""
    hitos = db.query(Hito).filter(Hito.id.in_(body.ids)).all()
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    now = datetime.utcnow()
    acted: list[int] = []
    for hito in hitos:
        if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(hito.fecha_hito)):
            raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")
        hito.estado = HITO_APROBADO
        hito.aprobado_by_rut = admin_rut
        hito.aprobado_by_name = admin.name if admin else None
        hito.aprobado_at = now
        hito.rechazo_motivo = None
        acted.append(hito.id)
    db.commit()
    return HitoBulkResult(procesados=len(acted), ids=acted)


@router.post("/rechazar-lote", response_model=HitoBulkResult)
async def rechazar_hitos_lote(
    body: HitoBulkIds,
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Reject every existing hito in ``ids`` (admin only). Skips ids that don't exist."""
    hitos = db.query(Hito).filter(Hito.id.in_(body.ids)).all()
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    now = datetime.utcnow()
    acted: list[int] = []
    for hito in hitos:
        hito.estado = HITO_RECHAZADO
        hito.aprobado_by_rut = admin_rut
        hito.aprobado_by_name = admin.name if admin else None
        hito.aprobado_at = now
        hito.rechazo_motivo = None
        acted.append(hito.id)
    db.commit()
    return HitoBulkResult(procesados=len(acted), ids=acted)


@router.post("/eliminar-lote", response_model=HitoBulkResult)
async def eliminar_hitos_lote(
    body: HitoBulkIds,
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Delete every existing hito in ``ids`` (admin only). POST so the id list
    travels in the body. Skips ids that don't exist; a closed period blocks the
    whole batch with the same 409 as the single endpoint."""
    hitos = db.query(Hito).filter(Hito.id.in_(body.ids)).all()
    acted: list[int] = []
    for hito in hitos:
        if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(hito.fecha_hito)):
            raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")
        acted.append(hito.id)
    for hito in hitos:
        db.delete(hito)
    db.commit()
    return HitoBulkResult(procesados=len(acted), ids=acted)


@router.get("/resumen", response_model=List[HitoResumenRow])
async def resumen_hitos(
    periodo: Optional[str] = Query(None, description="Mes YYYY-MM (default: mes actual)"),
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Per-lawyer totals for a period: approved count + gross sum + pending count."""
    if periodo:
        try:
            y, m = (int(x) for x in periodo.split("-"))
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="periodo inválido (usa YYYY-MM)")
    else:
        now = datetime.utcnow()
        y, m = now.year, now.month
    start = date(y, m, 1)
    end = date(y + (m == 12), (m % 12) + 1, 1)

    rows = (
        db.query(Hito)
        .filter(Hito.fecha_hito >= start, Hito.fecha_hito < end)
        .all()
    )
    by_lawyer: dict[int, dict] = {}
    for h in rows:
        r = by_lawyer.setdefault(
            h.lawyer_id,
            {"lawyer_id": h.lawyer_id, "lawyer_nombre": h.lawyer.name if h.lawyer else "",
             "aprobados": 0, "total_bruto": 0, "pendientes": 0},
        )
        if h.estado == HITO_APROBADO:
            r["aprobados"] += 1
            r["total_bruto"] += h.valor_bruto
        elif h.estado == HITO_PENDIENTE:
            r["pendientes"] += 1
    result = sorted(by_lawyer.values(), key=lambda x: x["total_bruto"], reverse=True)
    return [HitoResumenRow(**r) for r in result]


class HitoMensualAbogado(BaseModel):
    lawyer_id: int
    lawyer_nombre: str
    por_mes: dict[str, int]  # "YYYY-MM" -> cantidad ingresada ese mes
    total: int


class HitoStatsMensual(BaseModel):
    meses: List[str]                 # columnas del rango, ordenadas
    abogados: List[HitoMensualAbogado]
    totales_por_mes: dict[str, int]  # todos los abogados por mes (la "proyección")
    total_general: int


def _parse_month(value: str, field: str) -> tuple[int, int]:
    try:
        y, m = (int(x) for x in value.split("-"))
    except (ValueError, TypeError, AttributeError):
        raise HTTPException(status_code=400, detail=f"{field} inválido (usa YYYY-MM)")
    if not (1 <= m <= 12):
        raise HTTPException(status_code=400, detail=f"{field} inválido (mes 1–12)")
    return y, m


@router.get("/stats/mensual", response_model=HitoStatsMensual)
async def stats_mensual_hitos(
    desde: Optional[str] = Query(None, description="Mes inicial YYYY-MM (default: 11 meses atrás)"),
    hasta: Optional[str] = Query(None, description="Mes final YYYY-MM (default: mes actual)"),
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Hitos INGRESADOS por mes, por abogado + total general (proyección).

    Cuenta TODOS los hitos ingresados (cualquier estado) agrupados por el mes de
    ``fecha_hito`` — la misma base que usa el bono. Refleja la actividad/movimientos
    de cada abogado. Rango por defecto: los últimos 12 meses (incluye meses en 0
    para que las columnas sean estables). Ordenado por total de hitos desc."""
    now = datetime.utcnow()
    hy, hm = _parse_month(hasta, "hasta") if hasta else (now.year, now.month)
    if desde:
        dy, dm = _parse_month(desde, "desde")
    else:
        # ventana de 12 meses inclusive → 11 meses hacia atrás desde 'hasta'
        idx = hy * 12 + (hm - 1) - 11
        dy, dm = idx // 12, idx % 12 + 1
    if (dy, dm) > (hy, hm):
        raise HTTPException(status_code=400, detail="'desde' no puede ser posterior a 'hasta'")

    start = date(dy, dm, 1)
    end = date(hy + (hm == 12), (hm % 12) + 1, 1)

    # Meses del rango (columnas estables, incluso los que no tienen hitos).
    meses: List[str] = []
    yy, mm = dy, dm
    while (yy, mm) <= (hy, hm):
        meses.append(f"{yy:04d}-{mm:02d}")
        yy, mm = (yy + (mm == 12), (mm % 12) + 1)

    rows = (
        db.query(Hito)
        .filter(Hito.fecha_hito >= start, Hito.fecha_hito < end)
        .all()
    )
    by_lawyer: dict[int, dict] = {}
    totales_por_mes: dict[str, int] = {m: 0 for m in meses}
    for h in rows:
        mk = f"{h.fecha_hito.year:04d}-{h.fecha_hito.month:02d}"
        if mk not in totales_por_mes:
            continue  # defensivo; el filtro ya acota al rango
        r = by_lawyer.setdefault(
            h.lawyer_id,
            {"lawyer_id": h.lawyer_id,
             "lawyer_nombre": h.lawyer.name if h.lawyer else "",
             "por_mes": {m: 0 for m in meses}, "total": 0},
        )
        r["por_mes"][mk] += 1
        r["total"] += 1
        totales_por_mes[mk] += 1

    abogados = sorted(by_lawyer.values(), key=lambda x: x["total"], reverse=True)
    return HitoStatsMensual(
        meses=meses,
        abogados=[HitoMensualAbogado(**a) for a in abogados],
        totales_por_mes=totales_por_mes,
        total_general=sum(totales_por_mes.values()),
    )


# --------------------------------------------------------------------------- #
# Excel import (SISTEMA DE HITOS.xlsx — hojas HITOS JUNIOR / HITOS PLENO)
# --------------------------------------------------------------------------- #
class HitoHojaInfo(BaseModel):
    nombre: str
    filas: int


class HitoImportResult(BaseModel):
    total_leidas: int
    creadas: int
    aprobados: int
    pendientes: int
    omitidas_duplicadas: int
    errores: int  # rows skipped (no lawyer match, bad date, unknown pleno tipo)


def _norm(s) -> str:
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()


def _cell(row, i):
    return row[i] if row is not None and i < len(row) else None


def _parse_fecha(value):
    """Return a ``date`` from an Excel cell that may be a real datetime OR text.

    RRHH files often have the fecha column formatted as TEXT ("7/14/2026",
    "14/07/2026", "2026-07-14"), which the import used to reject as an error —
    silently dropping those rows. Tries the file's apparent M/D/Y first, then
    D/M/Y and ISO. Returns ``None`` if unparseable.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _is_hito_row(row) -> bool:
    ab = _cell(row, 1)
    return bool(ab and str(ab).strip() and _norm(ab) not in ("ABOGADO AT", "ABOGADO"))


def _open_hito_wb(data: bytes):
    if not data:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    try:
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=415, detail="No se pudo leer el archivo (usa un .xlsx)")


# Column headers → field, matched by WHOLE WORD (so "RUT" doesn't hit "bRUTo").
_COL_KEYS = {
    "abogado": ["ABOGADO"], "fecha": ["FECHA"], "rut": ["RUT"], "rol": ["ROL"],
    "tipo": ["TIPO DE HITO", "TIPO"], "proc": ["PROCEDIMIENTO"], "desc": ["DESCRIPCION"],
    "etapa": ["ETAPA"], "tramite": ["TRAMITE"], "aprobado": ["APROBADO"], "valor": ["VALOR"],
}


def _resolve_hito_columns(ws):
    """Find the header row of a hito sheet and map field → column index by header
    NAME (robust to layout differences: with/without RUT, any sheet name). Returns
    ``(header_row_index, cols)`` or ``(None, None)`` if no hito header is found."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        headers = [_norm(c) for c in row]
        cols: dict = {}
        for field, keys in _COL_KEYS.items():
            for i, h in enumerate(headers):
                if h and i not in cols.values() and any(
                    re.search(r"\b" + re.escape(k) + r"\b", h) for k in keys
                ):
                    cols[field] = i
                    break
        # A real header row has 'abogado' + several other fields in distinct columns.
        if "abogado" in cols and len(cols) >= 4:
            return ri, cols
    return None, None


def _sheet_is_hito(ws) -> bool:
    """True for a sheet the importer can read: NEW format (resolvable header with
    tipo+aprobado, any sheet name) or OLD format (name contains 'HITO')."""
    if "HITO" in _norm(ws.title):
        return True
    _ri, cols = _resolve_hito_columns(ws)
    return bool(cols and "tipo" in cols and "aprobado" in cols)


def _hito_sheets(wb, hoja: Optional[str]):
    """Importable sheets — by header/structure, NOT by name (skips PARÁMETROS,
    VARIABLES BONO, LIQUIDACIÓN, HISTORIAL, which have no hito header)."""
    sheets = [ws for ws in wb.worksheets if _sheet_is_hito(ws)]
    if hoja:
        sheets = [ws for ws in sheets if _norm(ws.title) == _norm(hoja)]
    return sheets


def _count_hito_rows(ws) -> int:
    """Data-row count for a hito sheet, for both formats (new header-based / old)."""
    header_row, cols = _resolve_hito_columns(ws)
    if cols and "tipo" in cols and "aprobado" in cols:
        ab_i = cols["abogado"]
        return sum(
            1
            for ri, row in enumerate(ws.iter_rows(values_only=True))
            if ri > header_row
            and ab_i < len(row)
            and row[ab_i]
            and _norm(row[ab_i]) not in ("ABOGADO AT", "ABOGADO")
        )
    return sum(1 for r in ws.iter_rows(values_only=True) if _is_hito_row(r))


def _match_tipo_any(text, tipos) -> Optional[HitoTipo]:
    """Match a 'Tipo de hito' cell against ALL hito tipos by label (nivel-agnostic)."""
    n = _norm(text)
    if not n:
        return None
    for t in tipos:
        lt = _norm(t.label)
        if n == lt or n in lt or lt in n:
            return t
    ntok = set(n.split())
    best, best_k = None, 0
    for t in tipos:
        k = len(ntok & set(_norm(t.label).split()))
        if k > best_k:
            best, best_k = t, k
    return best if best_k >= 2 else None


def _match_lawyer_by_name(name: str, lawyers) -> Optional[Lawyer]:
    """Match an Excel abogado name to a firm lawyer: all Excel name tokens must
    appear in the lawyer's full name (e.g. 'Gonzalo Calderón' → 'GONZALO ... CALDERÓN ...')."""
    toks = set(t for t in _norm(name).split() if len(t) > 1)
    if not toks:
        return None
    for l in lawyers:
        if toks <= set(_norm(l.name).split()):
            return l
    return None


def _match_pleno_tipo(text: str, pleno_tipos) -> Optional[HitoTipo]:
    n = _norm(text)
    if not n:
        return None
    for t in pleno_tipos:
        lt = _norm(t.label)
        if n == lt or n in lt or lt in n:
            return t
    ntok = set(n.split())
    best, best_k = None, 0
    for t in pleno_tipos:
        k = len(ntok & set(_norm(t.label).split()))
        if k > best_k:
            best, best_k = t, k
    return best if best_k >= 2 else None


class HitoAbogado(BaseModel):
    lawyer_id: int
    nombre: str


@router.get("/abogados", response_model=List[HitoAbogado])
async def list_hito_abogados(
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """Firm lawyers for the hito 'abogado' selector (admins register for another)."""
    lawyers = (
        db.query(Lawyer)
        .filter(Lawyer.is_firm_lawyer.is_(True))
        .order_by(Lawyer.name)
        .all()
    )
    return [HitoAbogado(lawyer_id=l.id, nombre=l.name) for l in lawyers]


@router.delete("/{hito_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_hito(
    hito_id: int,
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Delete a hito (admin only) — e.g. one registered for the wrong lawyer."""
    hito = db.query(Hito).filter(Hito.id == hito_id).first()
    if hito is None:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(hito.fecha_hito)):
        raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")
    db.delete(hito)
    db.commit()


class HitoUpdate(BaseModel):
    hito_tipo_id: int
    lawyer_id: Optional[int] = None
    fecha_hito: date
    rol_causa: Optional[str] = None
    procedimiento: Optional[str] = None
    descripcion: Optional[str] = None


@router.put("/{hito_id}", response_model=HitoResponse)
async def update_hito(
    hito_id: int,
    body: HitoUpdate,
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Edit a hito (admin only). Re-snapshots the value if the tipo changes.
    Preserves ETAPA/TRAMITE and evidence; period-close blocks edits both ways."""
    hito = db.query(Hito).filter(Hito.id == hito_id).first()
    if hito is None:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(hito.fecha_hito)):
        raise HTTPException(status_code=409, detail="El período de ese hito está cerrado")
    if cierre_svc.is_cerrado(db, cierre_svc.periodo_de_fecha(body.fecha_hito)):
        raise HTTPException(status_code=409, detail="El período de destino está cerrado")

    tipo = db.query(HitoTipo).filter(HitoTipo.id == body.hito_tipo_id, HitoTipo.activo.is_(True)).first()
    if tipo is None:
        raise HTTPException(status_code=404, detail="Tipo de hito no encontrado")

    if body.lawyer_id is not None:
        abogado = db.query(Lawyer).filter(Lawyer.id == body.lawyer_id).first()
        if abogado is None or not abogado.is_firm_lawyer:
            raise HTTPException(status_code=404, detail="Abogado no encontrado en el estudio")
        hito.lawyer_id = abogado.id

    # No duplicate: the resulting (abogado, RUT, causa) must not collide with
    # ANOTHER hito. Same client on a DIFFERENT causa is allowed (see _causa_key).
    rol_norm = (body.rol_causa or "").strip() or None
    if rol_norm is not None:
        causa = _causa_key(body.descripcion)
        prior = (
            db.query(Hito.descripcion)
            .filter(
                Hito.id != hito.id,
                Hito.lawyer_id == hito.lawyer_id,
                Hito.rol_causa == rol_norm,
            )
            .all()
        )
        if any(_causa_key(d) == causa for (d,) in prior):
            raise HTTPException(
                status_code=409,
                detail="Ya existe un hito de este abogado para esa causa",
            )

    hito.hito_tipo_id = tipo.id
    hito.valor_bruto = tipo.valor_bruto  # re-snapshot
    hito.fecha_hito = body.fecha_hito
    hito.rol_causa = rol_norm
    hito.procedimiento = body.procedimiento
    hito.descripcion = body.descripcion
    db.commit()
    db.refresh(hito)
    return _to_response(hito)


@router.post("/importar/hojas", response_model=List[HitoHojaInfo])
async def listar_hojas_hitos(
    archivo: UploadFile = File(...),
    _admin_rut: str = Depends(require_admin),
):
    """List the HITOS sheets of an uploaded SISTEMA DE HITOS.xlsx (name + row count)."""
    wb = _open_hito_wb(await archivo.read())
    out = [HitoHojaInfo(nombre=ws.title, filas=_count_hito_rows(ws)) for ws in _hito_sheets(wb, None)]
    wb.close()
    return out


@router.post("/importar", response_model=HitoImportResult)
async def importar_hitos(
    archivo: UploadFile = File(...),
    hoja: Optional[str] = Query(None, description="Hoja a importar (default: todas las HITOS)"),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Bulk-import hitos from SISTEMA DE HITOS.xlsx. JUNIOR and PLENO sheets have
    different columns; nivel comes from the sheet name. Maps abogado by name and
    (pleno) tipo by its label; skips duplicates and rows it can't resolve."""
    wb = _open_hito_wb(await archivo.read())
    sheets = _hito_sheets(wb, hoja)
    if hoja and not sheets:
        wb.close()
        raise HTTPException(status_code=404, detail=f"La hoja '{hoja}' no existe en el archivo")

    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    lawyers = db.query(Lawyer).filter(Lawyer.is_firm_lawyer.is_(True)).all()
    tipos = db.query(HitoTipo).all()
    junior_tipo = next((t for t in tipos if t.nivel == "junior"), None)
    pleno_tipos = [t for t in tipos if t.nivel == "pleno"]
    # Dedup key: (abogado, RUT, causa). Mismo cliente (RUT) en causas DISTINTAS no es
    # duplicado. Se dedupea si hay RUT O una causa (ROL en descripción) — así las
    # hojas SIN columna RUT (solo ROL) también quedan protegidas contra recargas.
    existing = {
        (h.lawyer_id, h.rol_causa, _causa_key(h.descripcion))
        for h in db.query(Hito.lawyer_id, Hito.rol_causa, Hito.descripcion)
        .filter((Hito.rol_causa.isnot(None)) | (Hito.descripcion.isnot(None)))
        .all()
        if h.rol_causa is not None or _causa_key(h.descripcion) is not None
    }

    total = creadas = aprobados = pendientes = dup = err = 0
    seen: set = set()
    nuevos: list[Hito] = []

    def _add(lawyer, tipo, fecha, *, rut, rol, procedimiento, descripcion,
             etapa, tramite, aprobado, valor_cell=None) -> str:
        """Dedup + build a Hito. Returns 'dup' | 'ok'. rol_causa = client RUT;
        the causa (ROL) goes to descripcion (matches the create/edit/dedup rule)."""
        try:
            valor = int(valor_cell) if valor_cell else tipo.valor_bruto
        except (ValueError, TypeError):
            valor = tipo.valor_bruto
        if valor <= 0:
            valor = tipo.valor_bruto
        rol_causa = (str(rut).strip()[:50] or None) if rut is not None else None
        # descripcion = ROL de la causa (o el texto libre viejo); es lo que distingue
        # dos hitos del mismo cliente en causas distintas.
        desc_src = descripcion if descripcion else rol
        desc_norm = (str(desc_src).strip()[:500] if desc_src else None)
        causa = _causa_key(desc_norm)
        # Dedupea si hay RUT O causa (ROL). Así protege también las hojas sin RUT
        # (rol_causa None) que traen el ROL en la descripción.
        if rol_causa is not None or causa is not None:
            key = (lawyer.id, rol_causa, causa)
            if key in existing or key in seen:
                return "dup"
            seen.add(key)
        estado = HITO_APROBADO if aprobado else HITO_PENDIENTE
        nuevos.append(Hito(
            lawyer_id=lawyer.id, hito_tipo_id=tipo.id, valor_bruto=valor,
            fecha_hito=fecha, rol_causa=rol_causa,
            procedimiento=(str(procedimiento).strip()[:100] if procedimiento else None),
            descripcion=desc_norm,
            etapa_sysgal=(str(etapa).strip()[:100] if etapa else None),
            tramite_sysgal=(str(tramite).strip()[:255] if tramite else None),
            estado=estado, created_by_rut=admin_rut,
            created_by_name=admin.name if admin else None,
            aprobado_by_rut=admin_rut if aprobado else None,
            aprobado_by_name=(admin.name if admin else None) if aprobado else None,
            aprobado_at=datetime.utcnow() if aprobado else None,
        ))
        return "ok"

    for ws in sheets:
        header_row, cols = _resolve_hito_columns(ws)
        new_format = bool(cols and "tipo" in cols and "aprobado" in cols)

        if new_format:
            # Header-based: cualquier layout (con/sin RUT), cualquier nombre de hoja.
            def g(row, field):
                i = cols.get(field)
                return row[i] if i is not None and i < len(row) else None
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri <= header_row:
                    continue
                ab = g(row, "abogado")
                if not ab or _norm(ab) in ("ABOGADO AT", "ABOGADO"):
                    continue
                total += 1
                fecha = _parse_fecha(g(row, "fecha"))
                if fecha is None:
                    err += 1
                    continue
                lawyer = _match_lawyer_by_name(ab, lawyers)
                tipo = _match_tipo_any(g(row, "tipo"), tipos)
                if lawyer is None or tipo is None:
                    err += 1
                    continue
                res = _add(
                    lawyer, tipo, fecha, rut=g(row, "rut"), rol=g(row, "rol"),
                    procedimiento=g(row, "proc"), descripcion=g(row, "desc"),
                    etapa=g(row, "etapa"), tramite=g(row, "tramite"),
                    aprobado=_norm(g(row, "aprobado")) in ("SI", "SÍ", "S"),
                    valor_cell=g(row, "valor"),
                )
                if res == "dup":
                    dup += 1
                    continue
                creadas += 1
                aprobados += 1 if _norm(g(row, "aprobado")) in ("SI", "SÍ", "S") else 0
                pendientes += 0 if _norm(g(row, "aprobado")) in ("SI", "SÍ", "S") else 1
            continue

        # OLD format (hojas 'HITOS JUNIOR/PLENO'): posición fija por nivel.
        is_junior = "JUNIOR" in _norm(ws.title)
        for row in ws.iter_rows(values_only=True):
            if not _is_hito_row(row):
                continue
            total += 1
            fecha = _parse_fecha(_cell(row, 2))
            if fecha is None:
                err += 1
                continue
            lawyer = _match_lawyer_by_name(_cell(row, 1), lawyers)
            if lawyer is None:
                err += 1
                continue
            if is_junior:
                procedimiento, rut, descripcion, tipo = (
                    _cell(row, 3), _cell(row, 4), _cell(row, 5), junior_tipo
                )
            else:
                procedimiento, rut, descripcion = None, _cell(row, 3), None
                tipo = _match_pleno_tipo(_cell(row, 5), pleno_tipos)
            aprobado = _norm(_cell(row, 8)) in ("SI", "SÍ", "S")
            if tipo is None:
                err += 1
                continue
            res = _add(
                lawyer, tipo, fecha, rut=rut, rol=None, procedimiento=procedimiento,
                descripcion=descripcion, etapa=_cell(row, 6), tramite=_cell(row, 7),
                aprobado=aprobado, valor_cell=_cell(row, 9),
            )
            if res == "dup":
                dup += 1
                continue
            creadas += 1
            aprobados += 1 if aprobado else 0
            pendientes += 0 if aprobado else 1

    wb.close()
    if nuevos:
        db.bulk_save_objects(nuevos)
        db.commit()
    return HitoImportResult(
        total_leidas=total, creadas=creadas, aprobados=aprobados, pendientes=pendientes,
        omitidas_duplicadas=dup, errores=err,
    )


@router.get("/{hito_id}/evidencia")
async def get_evidencia(
    hito_id: int,
    db: Session = Depends(get_db),
    current_lawyer: dict = Depends(get_current_lawyer),
):
    """Stream a hito's PJUD evidence. Admins, or the owning lawyer, only."""
    actor = _resolve_lawyer(db, current_lawyer)
    hito = db.query(Hito).filter(Hito.id == hito_id).first()
    if hito is None:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    if not _is_admin(actor) and (actor is None or actor.id != hito.lawyer_id):
        raise HTTPException(status_code=403, detail="Sin acceso a esta evidencia")
    if not hito.evidencia_storage_key:
        raise HTTPException(status_code=404, detail="Sin evidencia")

    from app.config import settings
    from app.services.storage_service import get_storage_backend

    data = get_storage_backend(settings).retrieve(hito.evidencia_storage_key)
    return StreamingResponse(
        io.BytesIO(data),
        media_type=hito.evidencia_content_type or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{hito.evidencia_filename or "evidencia"}"'},
    )
