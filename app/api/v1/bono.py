"""Bonus variables (V1–V4) + monthly liquidación — Hitos Slice 2.

Dirección Jurídica enters the manual monthly inputs (client/case counts,
complaints, renewals) per lawyer; the system computes V1–V4 and assembles the
liquidación (Fijo + Hitos aprobados + V1 + V3_neta + V2). All endpoints are
admin-only — payroll data. The math lives in ``app.services.bono_calc``.
"""
import io
import logging
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import get_db, require_admin
from app.models.bono import BonoVariables
from app.models.bono_cierre import BonoCierre, CIERRE_ABIERTO, CIERRE_CERRADO
from app.models.hito import Hito, HITO_APROBADO
from app.models.lawyer import Lawyer
from app.services import bono_calc
from app.services import bono_cierre_service as cierre_svc

logger = logging.getLogger(__name__)
router = APIRouter()

_NIVELES = {bono_calc.JUNIOR, bono_calc.PLENO}

# --- Activación AT import: name matching ----------------------------------- #
# The "Activación AT" .xlsx identifies lawyers by an external `codabo` + name,
# but our Lawyer model only has rut/name, so we match by name. Names in the
# report are reordered and unaccented vs the DB (e.g. "MELISSA BUSTOS ACEVEDO"
# vs "Melissa Denisse Acevedo Bustos"), so we compare normalized token SETS.

import re as _re
import unicodedata as _ud


def _norm_tokens(name: object) -> set:
    """Uppercase, accent-stripped token set of a name (drops 1-char tokens)."""
    s = _ud.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode().upper()
    return {t for t in _re.split(r"\s+", s) if len(t) > 1}


def _match_lawyer(nombre: object, candidates: list):
    """Best UNAMBIGUOUS lawyer match by shared-token count.

    ``candidates`` is a list of ``(lawyer, token_set)``. Requires >=2 shared
    tokens and no tie at the top score — otherwise returns ``None`` (reported
    as unmatched) so a fuzzy/ambiguous name is never silently mis-assigned.
    """
    ft = _norm_tokens(nombre)
    if not ft:
        return None
    scored = sorted(
        ((len(ft & toks), lw) for lw, toks in candidates), key=lambda x: x[0], reverse=True
    )
    if not scored or scored[0][0] < 2:
        return None
    if len(scored) > 1 and scored[1][0] == scored[0][0]:
        return None  # ambiguous: two lawyers tie on the top score
    return scored[0][1]


# --------------------------------------------------------------------------- #
# Schemas
# --------------------------------------------------------------------------- #
class BonoParametros(BaseModel):
    fijo_junior: int
    fijo_pleno: int
    v2_por_renovacion: int
    v4_pesos: dict  # {"leve":0.05,"medio":0.15,"grave":0.30}


class BonoVariablesIn(BaseModel):
    clientes_m2: int = 0
    clientes_activos: int = 0
    causas_asignadas: int = 0
    causas_cumplidas: int = 0
    # Avance de cartera por semana (puntos de %, ej. 7.5). Suman el % de V3.
    cumpl_sem1: float = 0.0
    cumpl_sem2: float = 0.0
    cumpl_sem3: float = 0.0
    cumpl_sem4: float = 0.0
    cumpl_sem5: float = 0.0
    meta_cumplimiento: float = 0.0
    reclamos_leve: int = 0
    reclamos_medio: int = 0
    reclamos_grave: int = 0
    renovaciones: int = 0
    verificado_dj: bool = False


class BonoRow(BaseModel):
    lawyer_id: int
    lawyer_nombre: str
    nivel: str
    has_row: bool  # whether variables were entered for this period
    verificado_dj: bool
    # manual inputs
    clientes_m2: int
    clientes_activos: int
    causas_asignadas: int
    causas_cumplidas: int
    cumpl_sem1: float = 0.0
    cumpl_sem2: float = 0.0
    cumpl_sem3: float = 0.0
    cumpl_sem4: float = 0.0
    cumpl_sem5: float = 0.0
    cumpl_total: float = 0.0  # suma de semanas (= % cumplimiento del mes)
    meta_cumplimiento: float = 0.0  # meta del mes (referencia; no altera V3)
    reclamos_leve: int
    reclamos_medio: int
    reclamos_grave: int
    renovaciones: int
    # computed breakdown
    fijo: int
    v1_pct_activacion: float
    v1_valor_cliente: int
    v1_bruto: int
    v3_pct_cumplimiento: float
    v3_tramo_bruto: int
    v4_pct: float
    v3_neta: int
    v2_bruto: int
    hitos_aprobados: int
    total_bono_gestion: int
    total_bruto: int


class BonoTotales(BaseModel):
    fijo: int
    hitos_aprobados: int
    v1_bruto: int
    v3_neta: int
    v2_bruto: int
    total_bono_gestion: int
    total_bruto: int


class LiquidacionResponse(BaseModel):
    periodo: str
    rows: List[BonoRow]
    totales: BonoTotales
    cerrado: bool = False
    cerrado_by: Optional[str] = None
    cerrado_at: Optional[datetime] = None
    sin_verificar: int = 0  # nivel lawyers without a "Verificado DJ" this period


class NivelIn(BaseModel):
    nivel: Optional[str] = None  # "junior" | "pleno" | null (removes from bonus)


class RosterRow(BaseModel):
    lawyer_id: int
    nombre: str
    nivel: Optional[str] = None
    is_active: bool


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _period_bounds(periodo: Optional[str]) -> tuple[str, date, date]:
    if periodo:
        try:
            y, m = (int(x) for x in periodo.split("-"))
            date(y, m, 1)  # validate
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="periodo inválido (usa YYYY-MM)")
    else:
        now = datetime.utcnow()
        y, m = now.year, now.month
    start = date(y, m, 1)
    end = date(y + (m == 12), (m % 12) + 1, 1)
    return f"{y:04d}-{m:02d}", start, end


def _renovaciones_map(db: Session, start: date, end: date) -> dict[int, int]:
    """Count of renovaciones per lawyer whose fecha_desde falls in the period.

    Feeds V2 directly from the Renovaciones module (no manual re-entry). NOTE: the
    Renovacion model has no 'primera cuota pagada' flag yet, so this counts every
    renovación of the month; the pago-condition would be a later addition."""
    from app.models.renovacion import Renovacion

    rows = (
        db.query(Renovacion.lawyer_id)
        .filter(
            Renovacion.lawyer_id.isnot(None),
            Renovacion.fecha_desde >= start,
            Renovacion.fecha_desde < end,
        )
        .all()
    )
    out: dict[int, int] = {}
    for (lid,) in rows:
        out[lid] = out.get(lid, 0) + 1
    return out


def _hitos_aprobados_map(db: Session, start: date, end: date) -> dict[int, int]:
    """Sum of approved hito gross value per lawyer for the period."""
    rows = (
        db.query(Hito)
        .filter(
            Hito.fecha_hito >= start,
            Hito.fecha_hito < end,
            Hito.estado == HITO_APROBADO,
        )
        .all()
    )
    out: dict[int, int] = {}
    for h in rows:
        out[h.lawyer_id] = out.get(h.lawyer_id, 0) + h.valor_bruto
    return out


def _build_row(
    lawyer: Lawyer, var: Optional[BonoVariables], hitos_aprobados: int,
    renovaciones_auto: Optional[int] = None,
) -> BonoRow:
    nivel = (var.nivel if var else None) or lawyer.nivel or bono_calc.JUNIOR
    # V2: renovaciones se cuentan automáticamente del módulo (renovaciones_auto);
    # el campo manual var.renovaciones queda como fallback si no se pasó el conteo.
    renovaciones = renovaciones_auto if renovaciones_auto is not None else (var.renovaciones if var else 0)
    inputs = dict(
        clientes_m2=var.clientes_m2 if var else 0,
        clientes_activos=var.clientes_activos if var else 0,
        causas_asignadas=var.causas_asignadas if var else 0,
        causas_cumplidas=var.causas_cumplidas if var else 0,
        reclamos_leve=var.reclamos_leve if var else 0,
        reclamos_medio=var.reclamos_medio if var else 0,
        reclamos_grave=var.reclamos_grave if var else 0,
        renovaciones=renovaciones,
    )
    semanas = [
        float(getattr(var, f"cumpl_sem{i}", 0) or 0) if var else 0.0 for i in range(1, 6)
    ]
    cumpl_total = round(sum(semanas), 4)
    b = bono_calc.compute(
        nivel, hitos_aprobados=hitos_aprobados, cumpl_semanas=semanas, **inputs
    )
    return BonoRow(
        lawyer_id=lawyer.id,
        lawyer_nombre=lawyer.name,
        nivel=nivel,
        has_row=var is not None,
        verificado_dj=bool(var.verificado_dj) if var else False,
        **inputs,
        cumpl_sem1=semanas[0],
        cumpl_sem2=semanas[1],
        cumpl_sem3=semanas[2],
        cumpl_sem4=semanas[3],
        cumpl_sem5=semanas[4],
        cumpl_total=cumpl_total,
        meta_cumplimiento=float(getattr(var, "meta_cumplimiento", 0) or 0) if var else 0.0,
        fijo=b["fijo"],
        v1_pct_activacion=b["v1_pct_activacion"],
        v1_valor_cliente=b["v1_valor_cliente"],
        v1_bruto=b["v1_bruto"],
        v3_pct_cumplimiento=b["v3_pct_cumplimiento"],
        v3_tramo_bruto=b["v3_tramo_bruto"],
        v4_pct=b["v4_pct"],
        v3_neta=b["v3_neta"],
        v2_bruto=b["v2_bruto"],
        hitos_aprobados=b["hitos_aprobados"],
        total_bono_gestion=b["total_bono_gestion"],
        total_bruto=b["total_bruto"],
    )


# --------------------------------------------------------------------------- #
# Excel export (RRHH)
# --------------------------------------------------------------------------- #
def _build_liquidacion_workbook(data: LiquidacionResponse):
    """Render the liquidación as an .xlsx mirroring the LIQUIDACIÓN MENSUAL sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    headers = [
        "Abogado", "Nivel", "Fijo mensual bruto", "Hitos H1 aprobados",
        "V1 Retención", "V3 neta (Cumpl.−Reclamos)", "V2 Renovación",
        "Total bono gestión bruto", "TOTAL BRUTO",
    ]
    money_cols = list(range(3, 10))  # C..I
    clp_fmt = '#,##0'

    ink = Font(name="Calibri", size=11)
    bold = Font(name="Calibri", size=11, bold=True)
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F2A44")
    total_fill = PatternFill("solid", fgColor="EEF1F6")
    thin = Side(style="thin", color="D6DAE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    # Title + note
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f"LIQUIDACIÓN MENSUAL · {data.periodo}")
    c.font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    note = ws.cell(
        row=2, column=1,
        value="Valores brutos (afectos a semana corrida). RRHH aplica SC en la liquidación final.",
    )
    note.font = Font(name="Calibri", size=10, italic=True, color="6B7280")

    header_row = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = white_bold
        cell.fill = head_fill
        cell.border = border
        cell.alignment = center if j <= 2 else right

    r = header_row + 1
    for row in data.rows:
        values = [
            row.lawyer_nombre, row.nivel, row.fijo, row.hitos_aprobados,
            row.v1_bruto, row.v3_neta, row.v2_bruto,
            row.total_bono_gestion, row.total_bruto,
        ]
        for j, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = ink
            cell.border = border
            if j in money_cols:
                cell.number_format = clp_fmt
                cell.alignment = right
            elif j == 2:
                cell.alignment = center
        r += 1

    # Total área
    t = data.totales
    total_values = [
        "TOTAL ÁREA", "", t.fijo, t.hitos_aprobados,
        t.v1_bruto, t.v3_neta, t.v2_bruto, t.total_bono_gestion, t.total_bruto,
    ]
    for j, v in enumerate(total_values, start=1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        if j in money_cols:
            cell.number_format = clp_fmt
            cell.alignment = right

    # Column widths
    widths = [30, 10, 18, 18, 16, 24, 16, 22, 18]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return wb


def _add_detalle_sheets(wb, db: Session, start: date, end: date, rows_by_lawyer: dict):
    """Add DETAIL sheets to the RRHH workbook so they see WHAT is being paid:
    'Detalle Hitos' (every approved hito) + 'Detalle Renovaciones' (each counted).

    ``rows_by_lawyer``: lawyer_id → BonoRow, only nivel lawyers (those in the bono).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.models.renovacion import Renovacion

    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F2A44")
    right = Alignment(horizontal="right")
    clp = "#,##0"

    def _sheet(title, headers, widths, money_cols):
        ws = wb.create_sheet(title)
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = white_bold
            c.fill = head_fill
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(row=2, column=1)
        return ws, money_cols

    lawyer_ids = list(rows_by_lawyer.keys())
    lawyer_name = {lid: r.lawyer_nombre for lid, r in rows_by_lawyer.items()}

    # --- Detalle Hitos: cada hito aprobado del período (lo que suma "Hitos H1") ---
    ws_h, hmoney = _sheet(
        "Detalle Hitos",
        ["Abogado", "Fecha", "Tipo de hito", "RUT/ROL", "Causa (ROL)", "Valor bruto"],
        [30, 12, 34, 16, 18, 14], {6},
    )
    hitos = (
        db.query(Hito)
        .filter(Hito.fecha_hito >= start, Hito.fecha_hito < end,
                Hito.estado == HITO_APROBADO, Hito.lawyer_id.in_(lawyer_ids))
        .order_by(Hito.lawyer_id, Hito.fecha_hito)
        .all()
    )
    r = 2
    for h in hitos:
        vals = [lawyer_name.get(h.lawyer_id, ""), h.fecha_hito,
                h.tipo.label if h.tipo else "", h.rol_causa or "",
                h.descripcion or "", h.valor_bruto]
        for j, v in enumerate(vals, start=1):
            cell = ws_h.cell(row=r, column=j, value=v)
            if j in hmoney:
                cell.number_format = clp
                cell.alignment = right
        r += 1

    # --- Resumen Renovaciones: cantidad por persona (TODAS las del período,
    #     incluidos procuradores y renovadores sin matchear a un abogado), tipo
    #     tabla dinámica, para reconciliar cuánto pagar de V2. ---
    all_renovs = (
        db.query(Renovacion)
        .filter(Renovacion.fecha_desde >= start, Renovacion.fecha_desde < end)
        .all()
    )
    matched_ids = {rn.lawyer_id for rn in all_renovs if rn.lawyer_id is not None}
    lw_info = {
        lw.id: lw.name + (" (proc.)" if lw.role == "procurador" else "")
        for lw in db.query(Lawyer).filter(Lawyer.id.in_(matched_ids)).all()
    } if matched_ids else {}
    counts: dict[str, int] = {}
    for rn in all_renovs:
        if rn.lawyer_id is not None:
            label = lw_info.get(rn.lawyer_id, f"#{rn.lawyer_id}")
        else:
            label = (rn.renovador_raw or "").strip() or "(sin asignar)"
        counts[label] = counts.get(label, 0) + 1

    ws_sum, _ = _sheet("Resumen Renovaciones", ["Abogado", "Cantidad"], [42, 12], set())
    total_font = Font(name="Calibri", size=11, bold=True)
    r = 2
    for label in sorted(counts, key=str.lower):
        ws_sum.cell(row=r, column=1, value=label)
        cell = ws_sum.cell(row=r, column=2, value=counts[label])
        cell.alignment = right
        r += 1
    lbl = ws_sum.cell(row=r, column=1, value="TOTAL")
    val = ws_sum.cell(row=r, column=2, value=sum(counts.values()))
    lbl.font = total_font
    val.font = total_font
    val.alignment = right

    # --- Detalle Renovaciones: cada renovación contada (lo que suma V2) ---
    ws_r, rmoney = _sheet(
        "Detalle Renovaciones",
        ["Abogado", "Fecha", "N° contrato", "Cliente", "RUT cliente", "Bono ($)"],
        [30, 12, 16, 30, 16, 12], {6},
    )
    renovs = (
        db.query(Renovacion)
        .filter(Renovacion.fecha_desde >= start, Renovacion.fecha_desde < end,
                Renovacion.lawyer_id.in_(lawyer_ids))
        .order_by(Renovacion.lawyer_id, Renovacion.fecha_desde)
        .all()
    )
    r = 2
    for rn in renovs:
        vals = [lawyer_name.get(rn.lawyer_id, ""), rn.fecha_desde, rn.numero_contrato,
                rn.cliente_nombre, rn.cliente_rut, bono_calc.V2_POR_RENOVACION]
        for j, v in enumerate(vals, start=1):
            cell = ws_r.cell(row=r, column=j, value=v)
            if j in rmoney:
                cell.number_format = clp
                cell.alignment = right
        r += 1


# --------------------------------------------------------------------------- #
# Endpoints
# --------------------------------------------------------------------------- #
@router.get("/parametros", response_model=BonoParametros)
async def get_parametros(_admin_rut: str = Depends(require_admin)):
    """The fixed bonus parameters, for display in the UI."""
    return BonoParametros(
        fijo_junior=bono_calc.FIJO[bono_calc.JUNIOR],
        fijo_pleno=bono_calc.FIJO[bono_calc.PLENO],
        v2_por_renovacion=bono_calc.V2_POR_RENOVACION,
        v4_pesos={
            "leve": bono_calc.V4_PESO_LEVE,
            "medio": bono_calc.V4_PESO_MEDIO,
            "grave": bono_calc.V4_PESO_GRAVE,
        },
    )


def _liquidacion_data(db: Session, periodo: Optional[str]) -> LiquidacionResponse:
    """Build the full liquidación (rows + totals) for a period. Shared by the
    JSON endpoint and the RRHH Excel export."""
    periodo, start, end = _period_bounds(periodo)
    lawyers = (
        db.query(Lawyer)
        .filter(Lawyer.nivel.in_(list(_NIVELES)))
        .order_by(Lawyer.nivel, Lawyer.name)
        .all()
    )
    hitos_map = _hitos_aprobados_map(db, start, end)
    renov_map = _renovaciones_map(db, start, end)
    vars_map = {
        v.lawyer_id: v
        for v in db.query(BonoVariables).filter(BonoVariables.periodo == periodo).all()
    }

    rows = [
        _build_row(lw, vars_map.get(lw.id), hitos_map.get(lw.id, 0), renov_map.get(lw.id, 0))
        for lw in lawyers
    ]
    totales = BonoTotales(
        fijo=sum(r.fijo for r in rows),
        hitos_aprobados=sum(r.hitos_aprobados for r in rows),
        v1_bruto=sum(r.v1_bruto for r in rows),
        v3_neta=sum(r.v3_neta for r in rows),
        v2_bruto=sum(r.v2_bruto for r in rows),
        total_bono_gestion=sum(r.total_bono_gestion for r in rows),
        total_bruto=sum(r.total_bruto for r in rows),
    )
    cierre = cierre_svc.get_cierre(db, periodo)
    cerrado = bool(cierre and cierre.estado == CIERRE_CERRADO)
    return LiquidacionResponse(
        periodo=periodo,
        rows=rows,
        totales=totales,
        cerrado=cerrado,
        cerrado_by=cierre.cerrado_by_name if cerrado else None,
        cerrado_at=cierre.cerrado_at if cerrado else None,
        sin_verificar=sum(1 for r in rows if not r.verificado_dj),
    )


@router.get("/liquidacion", response_model=LiquidacionResponse)
async def get_liquidacion(
    periodo: Optional[str] = Query(None, description="Mes YYYY-MM (default: mes actual)"),
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Full monthly liquidación for every bonus lawyer (nivel set), with totals."""
    return _liquidacion_data(db, periodo)


@router.get("/liquidacion/export")
async def export_liquidacion(
    periodo: Optional[str] = Query(None, description="Mes YYYY-MM (default: mes actual)"),
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Excel (.xlsx) of the monthly liquidación for RRHH to process payroll.

    Mirrors the firm's LIQUIDACIÓN MENSUAL sheet. All values are gross (afectos a
    semana corrida) — RRHH applies SC in their final liquidación.
    """
    per, start, end = _period_bounds(periodo)
    data = _liquidacion_data(db, periodo)
    wb = _build_liquidacion_workbook(data)
    _add_detalle_sheets(wb, db, start, end, {r.lawyer_id: r for r in data.rows})
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"liquidacion_{data.periodo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.put("/variables/{lawyer_id}", response_model=BonoRow)
async def upsert_variables(
    lawyer_id: int,
    body: BonoVariablesIn,
    periodo: str = Query(..., description="Mes YYYY-MM"),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Enter/update the manual bonus inputs for a lawyer/period (Dirección Jurídica)."""
    periodo, _start, _end = _period_bounds(periodo)
    if cierre_svc.is_cerrado(db, periodo):
        raise HTTPException(status_code=409, detail="El período está cerrado; reábrelo para editar")
    lawyer = db.query(Lawyer).filter(Lawyer.id == lawyer_id).first()
    if lawyer is None:
        raise HTTPException(status_code=404, detail="Abogado no encontrado")
    if lawyer.nivel not in _NIVELES:
        raise HTTPException(
            status_code=409,
            detail="El abogado no tiene nivel (junior/pleno) asignado para el bono",
        )
    for field in ("clientes_m2", "clientes_activos", "causas_asignadas", "causas_cumplidas",
                  "reclamos_leve", "reclamos_medio", "reclamos_grave", "renovaciones"):
        if getattr(body, field) < 0:
            raise HTTPException(status_code=400, detail=f"{field} no puede ser negativo")
    if body.clientes_activos > body.clientes_m2 and body.clientes_m2 > 0:
        raise HTTPException(status_code=400, detail="Clientes activos no puede superar a los M-2")
    if body.causas_cumplidas > body.causas_asignadas and body.causas_asignadas > 0:
        raise HTTPException(status_code=400, detail="Causas cumplidas no puede superar a las asignadas")
    # Avance semanal: cada semana entre 0 y 100 puntos de %; el total no supera 100.
    semanas = [body.cumpl_sem1, body.cumpl_sem2, body.cumpl_sem3, body.cumpl_sem4, body.cumpl_sem5]
    for i, s in enumerate(semanas, 1):
        if s < 0 or s > 100:
            raise HTTPException(status_code=400, detail=f"El % de la semana {i} debe estar entre 0 y 100")
    if sum(semanas) > 100.0001:
        raise HTTPException(status_code=400, detail="La suma de las semanas no puede superar 100%")
    if body.meta_cumplimiento < 0 or body.meta_cumplimiento > 100:
        raise HTTPException(status_code=400, detail="La meta debe estar entre 0 y 100%")

    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    var = (
        db.query(BonoVariables)
        .filter(BonoVariables.lawyer_id == lawyer_id, BonoVariables.periodo == periodo)
        .first()
    )
    if var is None:
        var = BonoVariables(
            lawyer_id=lawyer_id,
            periodo=periodo,
            nivel=lawyer.nivel,
            created_by_rut=admin_rut,
        )
        db.add(var)
    var.nivel = lawyer.nivel  # re-snapshot on each edit
    var.clientes_m2 = body.clientes_m2
    var.clientes_activos = body.clientes_activos
    var.causas_asignadas = body.causas_asignadas
    var.causas_cumplidas = body.causas_cumplidas
    var.cumpl_sem1 = body.cumpl_sem1
    var.cumpl_sem2 = body.cumpl_sem2
    var.cumpl_sem3 = body.cumpl_sem3
    var.cumpl_sem4 = body.cumpl_sem4
    var.cumpl_sem5 = body.cumpl_sem5
    var.meta_cumplimiento = body.meta_cumplimiento
    var.reclamos_leve = body.reclamos_leve
    var.reclamos_medio = body.reclamos_medio
    var.reclamos_grave = body.reclamos_grave
    var.renovaciones = body.renovaciones
    var.verificado_dj = body.verificado_dj
    var.updated_by_rut = admin_rut
    var.updated_by_name = admin.name if admin else None
    db.commit()
    db.refresh(var)

    _p, start, end = _period_bounds(periodo)
    hitos = _hitos_aprobados_map(db, start, end).get(lawyer_id, 0)
    renov = _renovaciones_map(db, start, end).get(lawyer_id, 0)
    return _build_row(lawyer, var, hitos, renov)


class ImportedRow(BaseModel):
    lawyer_id: int
    nombre: str
    clientes_m2: int
    clientes_activos: int
    v1_bruto: int


class ImportActivacionResult(BaseModel):
    updated: List[ImportedRow]        # lawyers whose V1 inputs were set
    skipped_no_nivel: List[str]       # matched, but no junior/pleno nivel → not in bono
    unmatched: List[str]              # file names with no confident lawyer match
    total_v1: int


@router.post("/variables/import-activacion", response_model=ImportActivacionResult)
async def import_activacion(
    periodo: str = Query(..., description="Mes YYYY-MM"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Bulk-load V1 inputs from a PJUD 'Activación AT' .xlsx.

    Sets ``clientes_m2`` (col ``totcontgral``) and ``clientes_activos`` (col
    ``tothiscartera``) per lawyer, matched by name. ONLY nivel-set lawyers are
    updated (matched Seniors without nivel are reported as skipped); every other
    bonus input on the row is preserved. Blocked when the period is closed.
    """
    periodo, start, end = _period_bounds(periodo)
    if cierre_svc.is_cerrado(db, periodo):
        raise HTTPException(status_code=409, detail="El período está cerrado; reábrelo para editar")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo (¿es un .xlsx?)")
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    try:
        i_nom = header.index("nomabo")
        i_f = header.index("totcontgral")
        i_i = header.index("tothiscartera")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="El archivo no parece un 'Activación AT' (faltan columnas nomabo/totcontgral/tothiscartera)",
        )

    firm = db.query(Lawyer).filter(Lawyer.is_firm_lawyer.is_(True)).all()
    candidates = [(lw, _norm_tokens(lw.name)) for lw in firm]
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()

    updated: List[ImportedRow] = []
    skipped: List[str] = []
    unmatched: List[str] = []
    for r in rows[1:]:
        nombre = r[i_nom] if i_nom < len(r) else None
        if not nombre:
            continue
        f_val = int(r[i_f] or 0) if i_f < len(r) else 0
        i_val = int(r[i_i] or 0) if i_i < len(r) else 0
        lw = _match_lawyer(nombre, candidates)
        if lw is None:
            unmatched.append(str(nombre))
            continue
        if lw.nivel not in _NIVELES:
            skipped.append(str(nombre))
            continue

        var = (
            db.query(BonoVariables)
            .filter(BonoVariables.lawyer_id == lw.id, BonoVariables.periodo == periodo)
            .first()
        )
        if var is None:
            var = BonoVariables(
                lawyer_id=lw.id, periodo=periodo, nivel=lw.nivel, created_by_rut=admin_rut
            )
            db.add(var)
        var.nivel = lw.nivel
        # Only the two V1 inputs are touched — every other input is preserved.
        var.clientes_m2 = f_val
        var.clientes_activos = i_val
        var.updated_by_rut = admin_rut
        var.updated_by_name = admin.name if admin else None

        v1 = bono_calc.compute(lw.nivel, clientes_m2=f_val, clientes_activos=i_val)["v1_bruto"]
        updated.append(
            ImportedRow(
                lawyer_id=lw.id, nombre=lw.name,
                clientes_m2=f_val, clientes_activos=i_val, v1_bruto=v1,
            )
        )

    db.commit()
    return ImportActivacionResult(
        updated=updated,
        skipped_no_nivel=skipped,
        unmatched=unmatched,
        total_v1=sum(u.v1_bruto for u in updated),
    )


@router.get("/roster", response_model=List[RosterRow])
async def get_roster(
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """All firm lawyers with their bonus nivel, to assign/adjust who is in the bono."""
    lawyers = (
        db.query(Lawyer)
        .filter(Lawyer.is_firm_lawyer.is_(True))
        .order_by(Lawyer.name)
        .all()
    )
    return [
        RosterRow(lawyer_id=l.id, nombre=l.name, nivel=l.nivel, is_active=bool(l.is_active))
        for l in lawyers
    ]


@router.put("/nivel/{lawyer_id}", response_model=dict)
async def set_nivel(
    lawyer_id: int,
    body: NivelIn,
    db: Session = Depends(get_db),
    _admin_rut: str = Depends(require_admin),
):
    """Assign/clear a lawyer's bonus tier (junior/pleno/null)."""
    if body.nivel is not None and body.nivel not in _NIVELES:
        raise HTTPException(status_code=400, detail="nivel debe ser 'junior', 'pleno' o null")
    lawyer = db.query(Lawyer).filter(Lawyer.id == lawyer_id).first()
    if lawyer is None:
        raise HTTPException(status_code=404, detail="Abogado no encontrado")
    lawyer.nivel = body.nivel
    db.commit()
    return {"lawyer_id": lawyer_id, "nivel": lawyer.nivel}


class CierreResponse(BaseModel):
    periodo: str
    cerrado: bool
    cerrado_by: Optional[str] = None
    cerrado_at: Optional[datetime] = None


@router.post("/cierre", response_model=CierreResponse)
async def cerrar_periodo(
    periodo: str = Query(..., description="Mes YYYY-MM"),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Close a bonus período — freezes its variables and hitos (payroll can't
    change after the fact). Idempotent: closing an already-closed period is a
    no-op. The caller (UI) warns about unverified rows before calling."""
    periodo, _s, _e = _period_bounds(periodo)
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    cierre = cierre_svc.get_cierre(db, periodo)
    if cierre is None:
        cierre = BonoCierre(periodo=periodo)
        db.add(cierre)
    cierre.estado = CIERRE_CERRADO
    cierre.cerrado_by_rut = admin_rut
    cierre.cerrado_by_name = admin.name if admin else None
    cierre.cerrado_at = datetime.utcnow()
    db.commit()
    db.refresh(cierre)
    return CierreResponse(
        periodo=periodo, cerrado=True,
        cerrado_by=cierre.cerrado_by_name, cerrado_at=cierre.cerrado_at,
    )


@router.post("/reabrir", response_model=CierreResponse)
async def reabrir_periodo(
    periodo: str = Query(..., description="Mes YYYY-MM"),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Reopen a closed período (admin), so its variables can be edited again."""
    periodo, _s, _e = _period_bounds(periodo)
    cierre = cierre_svc.get_cierre(db, periodo)
    if cierre is None or cierre.estado != CIERRE_CERRADO:
        raise HTTPException(status_code=409, detail="El período no está cerrado")
    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    cierre.estado = CIERRE_ABIERTO
    cierre.reabierto_by_rut = admin_rut
    cierre.reabierto_by_name = admin.name if admin else None
    cierre.reabierto_at = datetime.utcnow()
    db.commit()
    return CierreResponse(periodo=periodo, cerrado=False)
