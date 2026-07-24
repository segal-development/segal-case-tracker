"""Contract renewals (renovaciones) endpoints — new commercial module.

Minimal data entry (contract #, client RUT + name, handling lawyer, monthly
amount); everything else is derived: a renewal always runs 12 months, so
fecha_hasta = fecha_desde + 1 year, and total = monto_cuota * 12. The lawyer
selector is the firm's own active lawyers.
"""
import io
import logging
import unicodedata
from datetime import date, datetime
from typing import List, Optional

from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, field_validator
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.api.deps import get_current_lawyer, get_db, require_admin
from app.models.lawyer import Lawyer
from app.models.renovacion import Renovacion, CUOTAS_RENOVACION
from app.utils.rut import format_rut, normalize_rut

logger = logging.getLogger(__name__)
router = APIRouter()

MONTO_DEFAULT = 25_000


# --------------------------------------------------------------------------- #
# Schemas
# --------------------------------------------------------------------------- #
class RenovacionCreate(BaseModel):
    numero_contrato: str
    cliente_rut: str
    cliente_nombre: str
    lawyer_id: int
    monto_cuota: int = MONTO_DEFAULT
    fecha_desde: Optional[date] = None  # defaults to today

    @field_validator("numero_contrato", "cliente_rut", "cliente_nombre")
    @classmethod
    def _not_blank(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("campo obligatorio")
        return v.strip()


class RenovacionResponse(BaseModel):
    id: int
    numero_contrato: str
    cliente_rut: str  # formatted for display
    cliente_nombre: str
    lawyer_id: Optional[int] = None
    lawyer_nombre: Optional[str] = None
    renovador: Optional[str] = None  # display: system lawyer name or the raw name
    fecha_desde: date
    fecha_hasta: date
    monto_cuota: int
    cuotas: int
    total: int
    created_by_name: Optional[str] = None


class AbogadoOption(BaseModel):
    lawyer_id: int
    nombre: str


class RenovacionResumen(BaseModel):
    count: int
    total_cuotas: int  # sum of monthly cuotas (monthly recaudación base)
    total_anual: int   # sum of totals (cuota × 12)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _resolve_lawyer(db: Session, current_lawyer: dict) -> Optional[Lawyer]:
    sub = current_lawyer.get("sub") or current_lawyer.get("lawyer_id")
    if sub is None:
        return None
    if isinstance(sub, int) or (isinstance(sub, str) and str(sub).isdigit()):
        return db.query(Lawyer).filter(Lawyer.id == int(sub)).first()
    return db.query(Lawyer).filter(Lawyer.rut == str(sub)).first()


def _to_response(r: Renovacion) -> RenovacionResponse:
    return RenovacionResponse(
        id=r.id,
        numero_contrato=r.numero_contrato,
        cliente_rut=format_rut(r.cliente_rut) or r.cliente_rut,
        cliente_nombre=r.cliente_nombre,
        lawyer_id=r.lawyer_id,
        lawyer_nombre=r.lawyer.name if r.lawyer else None,
        renovador=(r.lawyer.name if r.lawyer else None) or r.renovador_raw,
        fecha_desde=r.fecha_desde,
        fecha_hasta=r.fecha_hasta,
        monto_cuota=r.monto_cuota,
        cuotas=r.cuotas,
        total=r.total,
        created_by_name=r.created_by_name,
    )


def _norm(s) -> str:
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()


def _build_renovador_index(db: Session) -> dict:
    """Map normalized 'INITIAL+SURNAME' usernames → firm lawyer id.

    e.g. 'Pablo Ismael Acevedo López' → {'PACEVEDO': id, 'PISMAEL': id, 'PLOPEZ': id};
    the Excel's 'PACEVEDO' then resolves to Pablo. First match wins on collisions.
    """
    idx: dict[str, int] = {}
    for l in db.query(Lawyer).filter(Lawyer.is_firm_lawyer.is_(True)).all():
        toks = _norm(l.name).split()
        if not toks:
            continue
        ini = toks[0][0]
        for tok in toks[1:]:
            idx.setdefault(ini + tok, l.id)
    return idx


def _match_renovador(renovador: Optional[str], idx: dict) -> Optional[int]:
    if not renovador:
        return None
    return idx.get(_norm(renovador))


def _period_bounds(periodo: Optional[str]) -> tuple[date, date]:
    if not periodo:
        now = datetime.utcnow()
        y, m = now.year, now.month
    else:
        try:
            y, m = (int(x) for x in periodo.split("-"))
            date(y, m, 1)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="periodo inválido (usa YYYY-MM)")
    return date(y, m, 1), date(y + (m == 12), (m % 12) + 1, 1)


# --------------------------------------------------------------------------- #
# Endpoints
# --------------------------------------------------------------------------- #
@router.get("/abogados", response_model=List[AbogadoOption])
async def list_abogados(
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """Active people selectable for a renovación: the firm's own lawyers PLUS
    procuradores (role='procurador'). Procuradores are not is_firm_lawyer, so
    they appear here but stay out of the Hitos/Bono/stats views."""
    lawyers = (
        db.query(Lawyer)
        .filter(
            Lawyer.is_active.is_(True),
            or_(Lawyer.is_firm_lawyer.is_(True), Lawyer.role == "procurador"),
        )
        .order_by(Lawyer.name)
        .all()
    )
    return [AbogadoOption(lawyer_id=l.id, nombre=l.name) for l in lawyers]


@router.post("", response_model=RenovacionResponse, status_code=status.HTTP_201_CREATED)
async def create_renovacion(
    body: RenovacionCreate,
    db: Session = Depends(get_db),
    current_lawyer: dict = Depends(get_current_lawyer),
):
    """Register a renewal. Derives fecha_hasta (+1 year) and total (cuota × 12)."""
    actor = _resolve_lawyer(db, current_lawyer)
    if actor is None:
        raise HTTPException(status_code=401, detail="No se pudo resolver el usuario")
    if body.monto_cuota <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")

    abogado = db.query(Lawyer).filter(Lawyer.id == body.lawyer_id).first()
    if abogado is None or not (abogado.is_firm_lawyer or abogado.role == "procurador"):
        raise HTTPException(status_code=404, detail="Abogado no encontrado en el estudio")

    desde = body.fecha_desde or date.today()
    hasta = desde + relativedelta(years=1)
    total = body.monto_cuota * CUOTAS_RENOVACION

    reno = Renovacion(
        numero_contrato=body.numero_contrato[:50],
        cliente_rut=normalize_rut(body.cliente_rut)[:20],
        cliente_nombre=body.cliente_nombre[:255],
        lawyer_id=abogado.id,
        renovador_raw=abogado.name[:100],
        fecha_desde=desde,
        fecha_hasta=hasta,
        monto_cuota=body.monto_cuota,
        cuotas=CUOTAS_RENOVACION,
        total=total,
        created_by_rut=actor.rut,
        created_by_name=actor.name,
    )
    db.add(reno)
    db.commit()
    db.refresh(reno)
    return _to_response(reno)


@router.get("", response_model=List[RenovacionResponse])
async def list_renovaciones(
    periodo: Optional[str] = Query(None, description="Filtrar por mes YYYY-MM (fecha de renovación)"),
    lawyer_id: Optional[int] = Query(None),
    q: Optional[str] = Query(None, description="Buscar por nombre, RUT o N° contrato"),
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """List renewals, most recent first, optionally filtered by month/lawyer/text."""
    query = db.query(Renovacion)
    if periodo:
        start, end = _period_bounds(periodo)
        query = query.filter(Renovacion.fecha_desde >= start, Renovacion.fecha_desde < end)
    if lawyer_id is not None:
        query = query.filter(Renovacion.lawyer_id == lawyer_id)
    if q and q.strip():
        like = f"%{q.strip()}%"
        query = query.filter(
            Renovacion.cliente_nombre.ilike(like)
            | Renovacion.cliente_rut.ilike(like)
            | Renovacion.numero_contrato.ilike(like)
        )
    rows = query.order_by(Renovacion.fecha_desde.desc(), Renovacion.id.desc()).all()
    return [_to_response(r) for r in rows]


@router.get("/resumen", response_model=RenovacionResumen)
async def resumen_renovaciones(
    periodo: Optional[str] = Query(None, description="Mes YYYY-MM (default: mes actual)"),
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """Count + recaudación totals for a month."""
    start, end = _period_bounds(periodo)
    rows = (
        db.query(Renovacion)
        .filter(Renovacion.fecha_desde >= start, Renovacion.fecha_desde < end)
        .all()
    )
    return RenovacionResumen(
        count=len(rows),
        total_cuotas=sum(r.monto_cuota for r in rows),
        total_anual=sum(r.total for r in rows),
    )


class ImportResult(BaseModel):
    total_leidas: int
    creadas: int
    vinculadas: int          # created linked to a system lawyer
    como_texto: int          # created with the renovador as text only
    omitidas_duplicadas: int
    errores: int


class HojaInfo(BaseModel):
    nombre: str
    filas: int


def _open_workbook(data: bytes):
    if not data:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    try:
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=415, detail="No se pudo leer el archivo (usa un .xlsx)")


def _is_data_row(row) -> bool:
    return bool(
        row and len(row) > 1 and row[1] and str(row[1]).strip() and _norm(row[1]) != "NOMBRE"
    )


def _importable_sheets(wb, hoja: Optional[str]):
    """Non-summary sheets, optionally narrowed to a single chosen sheet."""
    sheets = [ws for ws in wb.worksheets if "TOTAL" not in _norm(ws.title)]
    if hoja:
        sheets = [ws for ws in sheets if _norm(ws.title) == _norm(hoja)]
    return sheets


@router.post("/importar/hojas", response_model=List[HojaInfo])
async def listar_hojas(
    archivo: UploadFile = File(...),
    _admin_rut: str = Depends(require_admin),
):
    """List the importable sheets of an uploaded Excel (name + data-row count),
    so the UI can offer a per-sheet (e.g. per-year) selection before importing."""
    wb = _open_workbook(await archivo.read())
    out = []
    for ws in _importable_sheets(wb, None):
        filas = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if _is_data_row(row))
        out.append(HojaInfo(nombre=ws.title, filas=filas))
    wb.close()
    return out


@router.post("/importar", response_model=ImportResult)
async def importar_excel(
    archivo: UploadFile = File(...),
    hoja: Optional[str] = Query(None, description="Nombre de la hoja a importar (default: todas)"),
    db: Session = Depends(get_db),
    admin_rut: str = Depends(require_admin),
):
    """Bulk-import renovaciones from the firm's Excel. Maps each renovador to a
    system lawyer where possible (keeps the raw name otherwise), skips duplicates
    (same contrato + RUT + fecha), and reports a summary. If ``hoja`` is given,
    only that sheet is imported; otherwise all non-summary sheets."""
    wb = _open_workbook(await archivo.read())
    sheets = _importable_sheets(wb, hoja)
    if hoja and not sheets:
        wb.close()
        raise HTTPException(status_code=404, detail=f"La hoja '{hoja}' no existe en el archivo")

    admin = db.query(Lawyer).filter(Lawyer.rut == admin_rut).first()
    idx = _build_renovador_index(db)
    existing = {
        (r.numero_contrato, r.cliente_rut, r.fecha_desde)
        for r in db.query(
            Renovacion.numero_contrato, Renovacion.cliente_rut, Renovacion.fecha_desde
        ).all()
    }

    total = creadas = vinculadas = como_texto = dup = err = 0
    seen: set = set()
    nuevos: list[Renovacion] = []

    for ws in sheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8:
                continue
            nombre = row[1]
            if not nombre or not str(nombre).strip() or _norm(nombre) == "NOMBRE":
                continue
            total += 1
            desde = row[4]
            if not isinstance(desde, datetime):
                err += 1
                continue
            desde = desde.date()
            # Truncate to column limits — real files have junk (e.g. a name in
            # the RUT cell) that Postgres would reject; never 500 on one bad cell.
            contrato = (str(row[2]).strip() if row[2] is not None else "")[:50]
            if not contrato:
                err += 1
                continue
            rut = (normalize_rut(str(row[0])) if row[0] is not None else "")[:20]
            try:
                cuotas = int(row[3]) if row[3] else CUOTAS_RENOVACION
            except (ValueError, TypeError):
                cuotas = CUOTAS_RENOVACION
            if cuotas <= 0 or cuotas > 60:
                cuotas = CUOTAS_RENOVACION
            hasta = row[5].date() if isinstance(row[5], datetime) else desde + relativedelta(years=1)
            try:
                monto = int(row[7]) if row[7] else MONTO_DEFAULT
            except (ValueError, TypeError):
                monto = MONTO_DEFAULT
            if monto <= 0 or monto > 1_000_000:
                monto = MONTO_DEFAULT
            renovador = (str(row[6]).strip()[:100] if row[6] else None)
            lawyer_id = _match_renovador(renovador, idx)

            key = (contrato, rut, desde)
            if key in existing or key in seen:
                dup += 1
                continue
            seen.add(key)
            nuevos.append(Renovacion(
                numero_contrato=contrato, cliente_rut=rut, cliente_nombre=str(nombre).strip()[:255],
                lawyer_id=lawyer_id, renovador_raw=renovador,
                fecha_desde=desde, fecha_hasta=hasta,
                monto_cuota=monto, cuotas=cuotas, total=monto * cuotas,
                created_by_rut=admin_rut, created_by_name=admin.name if admin else None,
            ))
            creadas += 1
            if lawyer_id:
                vinculadas += 1
            else:
                como_texto += 1

    wb.close()
    if nuevos:
        db.bulk_save_objects(nuevos)
        db.commit()
    return ImportResult(
        total_leidas=total, creadas=creadas, vinculadas=vinculadas,
        como_texto=como_texto, omitidas_duplicadas=dup, errores=err,
    )


class RecaudacionMes(BaseModel):
    mes: int
    count: int
    recaudacion: int         # sum of monthly cuotas started that month
    proyeccion_anual: int    # recaudacion × 12


class RecaudacionAnual(BaseModel):
    anio: int
    meses: List[RecaudacionMes]
    total_count: int
    total_recaudacion: int
    total_proyeccion: int


@router.get("/recaudacion", response_model=RecaudacionAnual)
async def recaudacion(
    anio: Optional[int] = Query(None, description="Año (default: actual)"),
    db: Session = Depends(get_db),
    _lawyer: dict = Depends(get_current_lawyer),
):
    """Monthly + annual recaudación for a year (mirrors the TOTALES sheet)."""
    if anio is None:
        anio = datetime.utcnow().year
    rows = (
        db.query(Renovacion)
        .filter(Renovacion.fecha_desde >= date(anio, 1, 1), Renovacion.fecha_desde < date(anio + 1, 1, 1))
        .all()
    )
    agg = {m: [0, 0] for m in range(1, 13)}  # mes → [count, recaudacion]
    for r in rows:
        a = agg[r.fecha_desde.month]
        a[0] += 1
        a[1] += r.monto_cuota
    meses = [
        RecaudacionMes(mes=m, count=agg[m][0], recaudacion=agg[m][1], proyeccion_anual=agg[m][1] * 12)
        for m in range(1, 13)
    ]
    return RecaudacionAnual(
        anio=anio,
        meses=meses,
        total_count=sum(x.count for x in meses),
        total_recaudacion=sum(x.recaudacion for x in meses),
        total_proyeccion=sum(x.proyeccion_anual for x in meses),
    )


@router.delete("/{renovacion_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_renovacion(
    renovacion_id: int,
    db: Session = Depends(get_db),
    current_lawyer: dict = Depends(get_current_lawyer),
):
    """Delete a renewal (admin, or the person who registered it)."""
    actor = _resolve_lawyer(db, current_lawyer)
    reno = db.query(Renovacion).filter(Renovacion.id == renovacion_id).first()
    if reno is None:
        raise HTTPException(status_code=404, detail="Renovación no encontrada")
    is_admin = bool(actor and actor.role == "admin")
    if not is_admin and (actor is None or reno.created_by_rut != actor.rut):
        raise HTTPException(status_code=403, detail="Sin permiso para eliminar esta renovación")
    db.delete(reno)
    db.commit()
