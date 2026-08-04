"""Tests for the bonus variables + liquidación endpoints (Hitos Slice 2).

Covers admin-only access, nivel assignment, the upsert of manual inputs with its
validations, and the assembled liquidación (Fijo + Hitos aprobados + V1+V3+V2).
"""
from datetime import date

import pytest

from app.core.security import create_access_token
from app.models.hito import Hito, HitoTipo, HITO_APROBADO, HITO_PENDIENTE
from app.models.lawyer import Lawyer

ADMIN_RUT = "16021492-9"
LAWYER_RUT = "19643548-4"


@pytest.fixture
def admin(db):
    obj = Lawyer(rut=ADMIN_RUT, name="Carla Admin", role="admin")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@pytest.fixture
def junior(db):
    obj = Lawyer(rut=LAWYER_RUT, name="Fernanda Arroyo", role="lawyer", nivel="junior")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def _h(rut):
    return {"Authorization": "Bearer " + create_access_token({"sub": rut})}


def test_requires_admin(client, junior):
    r = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(LAWYER_RUT))
    assert r.status_code == 403


def test_parametros(client, admin):
    r = client.get("/api/v1/bono/parametros", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    body = r.json()
    assert body["fijo_junior"] == 1_105_354
    assert body["fijo_pleno"] == 1_363_354
    assert body["v2_por_renovacion"] == 10_400


def test_set_nivel_and_liquidacion_includes_only_nivel_lawyers(client, admin, db):
    # a lawyer with no nivel is excluded until assigned
    lw = Lawyer(rut="11111111-1", name="Sandy Quijada", role="lawyer")
    db.add(lw)
    db.commit()
    db.refresh(lw)

    r = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert all(row["lawyer_id"] != lw.id for row in r.json()["rows"])

    r = client.put(f"/api/v1/bono/nivel/{lw.id}", headers=_h(ADMIN_RUT), json={"nivel": "pleno"})
    assert r.status_code == 200
    r = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT))
    row = next(x for x in r.json()["rows"] if x["lawyer_id"] == lw.id)
    assert row["nivel"] == "pleno"
    assert row["fijo"] == 1_363_354


def test_roster_lists_firm_lawyers_with_nivel(client, admin, junior, db):
    extra = Lawyer(rut="33333333-3", name="Aaa Sin Nivel", role="lawyer")  # firm lawyer, no nivel
    db.add(extra)
    db.commit()
    r = client.get("/api/v1/bono/roster", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    body = {row["lawyer_id"]: row for row in r.json()}
    assert body[junior.id]["nivel"] == "junior"
    assert body[extra.id]["nivel"] is None  # assignable from the UI


def test_set_nivel_rejects_bad_value(client, admin, junior):
    r = client.put(f"/api/v1/bono/nivel/{junior.id}", headers=_h(ADMIN_RUT), json={"nivel": "senior"})
    assert r.status_code == 400


def test_upsert_variables_computes_bonus(client, admin, junior):
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={
            "clientes_m2": 100, "clientes_activos": 90,
            "causas_asignadas": 100, "causas_cumplidas": 95,
            "renovaciones": 2, "verificado_dj": True,
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["v1_bruto"] == 8000 * 90
    assert body["v3_neta"] == 100_000
    assert body["v2_bruto"] == 0  # V2 ahora se cuenta del módulo de renovaciones (0 acá)
    assert body["has_row"] is True
    assert body["verificado_dj"] is True

    # upsert is idempotent on the same period (updates, not duplicates)
    r2 = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={"clientes_m2": 100, "clientes_activos": 80},
    )
    assert r2.json()["v1_bruto"] == 8000 * 80  # 80% still ≥80% tramo
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    assert sum(1 for row in liq["rows"] if row["lawyer_id"] == junior.id) == 1


def test_upsert_rejects_activos_gt_m2(client, admin, junior):
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={"clientes_m2": 50, "clientes_activos": 80},
    )
    assert r.status_code == 400


def test_upsert_rejects_lawyer_without_nivel(client, admin, db):
    lw = Lawyer(rut="22222222-2", name="Sin Nivel", role="lawyer")
    db.add(lw)
    db.commit()
    db.refresh(lw)
    r = client.put(
        f"/api/v1/bono/variables/{lw.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={"clientes_m2": 10, "clientes_activos": 5},
    )
    assert r.status_code == 409


def test_export_liquidacion_xlsx(client, admin, junior):
    client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={"clientes_m2": 100, "clientes_activos": 90, "renovaciones": 2},
    )
    r = client.get("/api/v1/bono/liquidacion/export?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "liquidacion_2026-07.xlsx" in r.headers["content-disposition"]

    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "Fernanda Arroyo" in flat
    assert 8000 * 90 in flat  # V1 bruto lands in a cell
    assert any(str(v).startswith("TOTAL ÁREA") for v in flat if v)


def test_export_requires_admin(client, junior):
    r = client.get("/api/v1/bono/liquidacion/export?periodo=2026-07", headers=_h(LAWYER_RUT))
    assert r.status_code == 403


def test_cierre_freezes_variables_and_reabrir_unlocks(client, admin, junior):
    # can edit while open
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT), json={"clientes_m2": 100, "clientes_activos": 90},
    )
    assert r.status_code == 200

    # close the period
    c = client.post("/api/v1/bono/cierre?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert c.status_code == 200
    assert c.json()["cerrado"] is True

    # liquidación reports it closed
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    assert liq["cerrado"] is True
    assert liq["cerrado_by"] == "Carla Admin"

    # editing a closed period is blocked
    blocked = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT), json={"clientes_m2": 50, "clientes_activos": 10},
    )
    assert blocked.status_code == 409

    # reopen → editable again
    ro = client.post("/api/v1/bono/reabrir?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert ro.status_code == 200
    assert ro.json()["cerrado"] is False
    again = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT), json={"clientes_m2": 50, "clientes_activos": 10},
    )
    assert again.status_code == 200


def test_closed_period_blocks_hito_approval(client, admin, junior, db):
    tipo = HitoTipo(code="j_p", label="Presc", nivel="junior", valor_bruto=8_077, orden=1)
    db.add(tipo)
    db.commit()
    db.refresh(tipo)
    h = Hito(
        lawyer_id=junior.id, hito_tipo_id=tipo.id, valor_bruto=8_077,
        fecha_hito=date(2026, 7, 15), estado=HITO_PENDIENTE, evidencia_storage_key="x",
    )
    db.add(h)
    db.commit()
    db.refresh(h)

    client.post("/api/v1/bono/cierre?periodo=2026-07", headers=_h(ADMIN_RUT))
    r = client.post(f"/api/v1/hitos/{h.id}/aprobar", headers=_h(ADMIN_RUT))
    assert r.status_code == 409
    assert "cerrado" in r.json()["detail"].lower()


def test_reabrir_open_period_is_409(client, admin):
    r = client.post("/api/v1/bono/reabrir?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 409


def test_cierre_requires_admin(client, junior):
    r = client.post("/api/v1/bono/cierre?periodo=2026-07", headers=_h(LAWYER_RUT))
    assert r.status_code == 403


def test_sin_verificar_count(client, admin, junior):
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-08", headers=_h(ADMIN_RUT)).json()
    # junior has no row for 2026-08 → not verified → counted
    assert liq["sin_verificar"] >= 1


def test_liquidacion_includes_approved_hitos(client, admin, junior, db):
    tipo = HitoTipo(code="j_presc", label="Prescripción", nivel="junior", valor_bruto=8_077, orden=1)
    db.add(tipo)
    db.commit()
    db.refresh(tipo)
    h = Hito(
        lawyer_id=junior.id, hito_tipo_id=tipo.id, valor_bruto=8_077,
        fecha_hito=date(2026, 7, 15), estado=HITO_APROBADO,
        evidencia_storage_key="x",
    )
    db.add(h)
    db.commit()

    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["hitos_aprobados"] == 8_077
    assert row["total_bruto"] == row["fijo"] + 8_077 + row["total_bono_gestion"]


# --- Import Activación AT ---------------------------------------------------- #
import io

from openpyxl import Workbook

from app.api.v1.bono import _match_lawyer, _norm_tokens


def _acti_xlsx(rows):
    """Build an in-memory 'Activación AT' .xlsx. rows: (nomabo, totcontgral, tothiscartera).

    Columns are located by NAME in the endpoint, so order here is irrelevant as
    long as the three named headers exist.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["codabo", "nomabo", "totcontcartera", "totcontgral", "tothiscartera"])
    for nom, f, i in rows:
        ws.append([1, nom, 0, f, i])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _post_acti(client, rows, periodo="2026-07"):
    return client.post(
        f"/api/v1/bono/variables/import-activacion?periodo={periodo}",
        headers=_h(ADMIN_RUT),
        files={"file": ("acti.xlsx", _acti_xlsx(rows),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_match_lawyer_reordered_and_accented():
    fern = Lawyer(rut="1-9", name="Fernanda Arroyo Diaz", nivel="junior")
    pablo = Lawyer(rut="2-7", name="Pablo Ismael Acevedo López")
    cands = [(fern, _norm_tokens(fern.name)), (pablo, _norm_tokens(pablo.name))]
    # reordered + missing accent in the file name still matches
    assert _match_lawyer("PABLO LOPEZ ACEVEDO", cands) is pablo
    # only one shared token → ambiguous/insufficient → no match
    assert _match_lawyer("PABLO SOLANO", cands) is None


def test_import_sets_v1_inputs_and_computes(client, admin, junior):
    # junior fixture = "Fernanda Arroyo". 34/40 = 85% → $8.000 × 34 = 272_000
    r = _post_acti(client, [("FERNANDA ARROYO DIAZ", 40, 34)])
    assert r.status_code == 200
    body = r.json()
    assert len(body["updated"]) == 1
    up = body["updated"][0]
    assert up["clientes_m2"] == 40 and up["clientes_activos"] == 34
    assert up["v1_bruto"] == 8000 * 34
    assert body["total_v1"] == 8000 * 34
    # and it lands in the liquidación
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["v1_bruto"] == 8000 * 34


def test_import_preserves_other_inputs(client, admin, junior):
    # pre-set el avance semanal (V3) vía el upsert normal
    client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        json={"cumpl_sem1": 50, "cumpl_sem2": 45},
    )
    _post_acti(client, [("FERNANDA ARROYO DIAZ", 40, 34)])
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["v1_bruto"] == 8000 * 34   # import setea V1
    assert row["cumpl_total"] == 95       # y NO pisa el avance semanal previo


def test_import_skips_matched_lawyer_without_nivel(client, admin, db):
    senior = Lawyer(rut="55555555-5", name="Melissa Denisse Acevedo Bustos", role="lawyer")
    db.add(senior)
    db.commit()
    r = _post_acti(client, [("MELISSA BUSTOS ACEVEDO", 45, 34)])
    body = r.json()
    assert body["updated"] == []
    assert "MELISSA BUSTOS ACEVEDO" in body["skipped_no_nivel"]


def test_import_reports_unmatched(client, admin, junior):
    r = _post_acti(client, [("QUIEN SABE PERSONA", 10, 8)])
    body = r.json()
    assert body["updated"] == []
    assert "QUIEN SABE PERSONA" in body["unmatched"]


def test_import_blocked_when_period_closed(client, admin, junior):
    client.post("/api/v1/bono/cierre?periodo=2026-07", headers=_h(ADMIN_RUT))
    r = _post_acti(client, [("FERNANDA ARROYO DIAZ", 40, 34)])
    assert r.status_code == 409


def test_import_rejects_wrong_columns(client, admin):
    wb = Workbook(); ws = wb.active
    ws.append(["foo", "bar", "baz"]); ws.append([1, 2, 3])
    buf = io.BytesIO(); wb.save(buf)
    r = client.post(
        "/api/v1/bono/variables/import-activacion?periodo=2026-07",
        headers=_h(ADMIN_RUT),
        files={"file": ("bad.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400


def test_import_requires_admin(client, junior):
    r = client.post(
        "/api/v1/bono/variables/import-activacion?periodo=2026-07",
        headers=_h(LAWYER_RUT),
        files={"file": ("acti.xlsx", _acti_xlsx([("FERNANDA ARROYO DIAZ", 40, 34)]),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 403


def test_avance_semanal_alimenta_v3(client, admin, junior):
    """Carla carga % por semana → suman el cumplimiento → V3 refleja el tramo."""
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-08",
        headers=_h(ADMIN_RUT),
        json={"cumpl_sem1": 30, "cumpl_sem2": 30, "cumpl_sem3": 20, "cumpl_sem4": 12},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["cumpl_total"] == 92          # suma de semanas
    assert body["v3_pct_cumplimiento"] == 0.92
    assert body["v3_neta"] == 100_000         # junior ≥90%
    # persistió: reaparece en la liquidación
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-08", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["cumpl_sem2"] == 30 and row["v3_neta"] == 100_000


def test_avance_semanal_rechaza_suma_mayor_a_100(client, admin, junior):
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-08",
        headers=_h(ADMIN_RUT),
        json={"cumpl_sem1": 60, "cumpl_sem2": 60},
    )
    assert r.status_code == 400


def test_meta_cumplimiento_se_guarda_y_expone(client, admin, junior):
    """Carla carga la meta del mes a mano; se guarda y aparece junto a lo que llevan."""
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-08",
        headers=_h(ADMIN_RUT),
        json={"cumpl_sem1": 30, "cumpl_sem2": 25, "meta_cumplimiento": 90},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["cumpl_total"] == 55        # lo que llevan
    assert body["meta_cumplimiento"] == 90  # la meta cargada
    # la meta NO altera V3 (55% < 80% → sigue $0)
    assert body["v3_neta"] == 0
    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-08", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["meta_cumplimiento"] == 90


def test_meta_fuera_de_rango_rechaza(client, admin, junior):
    r = client.put(
        f"/api/v1/bono/variables/{junior.id}?periodo=2026-08",
        headers=_h(ADMIN_RUT),
        json={"meta_cumplimiento": 150},
    )
    assert r.status_code == 400


def test_v2_cuenta_renovaciones_del_modulo(client, db, admin, junior):
    """V2 se cuenta automáticamente de las renovaciones del módulo (no del campo manual)."""
    from datetime import date as _date
    from app.models.renovacion import Renovacion
    for i in range(3):  # 3 renovaciones de julio para este abogado
        db.add(Renovacion(
            numero_contrato=f"C{i}", cliente_rut=f"1111111{i}-1", cliente_nombre="X",
            lawyer_id=junior.id, fecha_desde=_date(2026, 7, 10), fecha_hasta=_date(2027, 7, 10),
            monto_cuota=25000, cuotas=12, total=300000,
        ))
    # una de OTRO mes → no cuenta
    db.add(Renovacion(numero_contrato="Cjun", cliente_rut="9999999-9", cliente_nombre="Y",
                      lawyer_id=junior.id, fecha_desde=_date(2026, 6, 1), fecha_hasta=_date(2027, 6, 1),
                      monto_cuota=25000, cuotas=12, total=300000))
    db.commit()

    liq = client.get("/api/v1/bono/liquidacion?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    row = next(x for x in liq["rows"] if x["lawyer_id"] == junior.id)
    assert row["renovaciones"] == 3               # solo las de julio
    assert row["v2_bruto"] == 3 * 10_400          # V2 automático


def test_export_incluye_hojas_de_detalle(client, db, admin, junior):
    """El Excel de RRHH trae hojas aparte con el detalle: hitos + renovaciones."""
    from datetime import date as _date
    from app.models.renovacion import Renovacion
    tipo = HitoTipo(code="j_det", label="Conversión preventiva", nivel="junior", valor_bruto=3000, orden=1)
    db.add(tipo); db.commit(); db.refresh(tipo)
    db.add(Hito(lawyer_id=junior.id, hito_tipo_id=tipo.id, valor_bruto=3000,
                fecha_hito=_date(2026, 7, 5), estado=HITO_APROBADO,
                rol_causa="12.345.678-5", descripcion="C-100-2026"))
    db.add(Renovacion(numero_contrato="R1", cliente_rut="9.999.999-9", cliente_nombre="Cliente X",
                      lawyer_id=junior.id, fecha_desde=_date(2026, 7, 8), fecha_hasta=_date(2027, 7, 8),
                      monto_cuota=25000, cuotas=12, total=300000))
    db.commit()

    r = client.get("/api/v1/bono/liquidacion/export?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "Detalle Hitos" in wb.sheetnames
    assert "Detalle Renovaciones" in wb.sheetnames
    # el hito aparece en su hoja
    hflat = [c.value for row in wb["Detalle Hitos"].iter_rows() for c in row]
    assert "C-100-2026" in hflat and 3000 in hflat
    # la renovación aparece con su bono
    rflat = [c.value for row in wb["Detalle Renovaciones"].iter_rows() for c in row]
    assert "Cliente X" in rflat and 10_400 in rflat


def test_export_resumen_renovaciones_cuenta_todos(client, db, admin, junior):
    """El resumen (tipo tabla dinámica) cuenta renovaciones por persona, incluidos
    los renovadores sin matchear a un abogado, con fila TOTAL."""
    from datetime import date as _date
    from app.models.renovacion import Renovacion
    for i in range(2):  # 2 matcheadas al abogado
        db.add(Renovacion(numero_contrato=f"M{i}", cliente_rut=f"1010101{i}-1", cliente_nombre="X",
                          lawyer_id=junior.id, fecha_desde=_date(2026, 7, 3), fecha_hasta=_date(2027, 7, 3),
                          monto_cuota=25000, cuotas=12, total=300000))
    # una SIN abogado (renovador_raw crudo) → igual aparece en el resumen
    db.add(Renovacion(numero_contrato="U1", cliente_rut="2020202-2", cliente_nombre="Y",
                      lawyer_id=None, renovador_raw="CCARO", fecha_desde=_date(2026, 7, 4),
                      fecha_hasta=_date(2027, 7, 4), monto_cuota=25000, cuotas=12, total=300000))
    # una de junio → no cuenta
    db.add(Renovacion(numero_contrato="Jun", cliente_rut="3030303-3", cliente_nombre="Z",
                      lawyer_id=junior.id, fecha_desde=_date(2026, 6, 1), fecha_hasta=_date(2027, 6, 1),
                      monto_cuota=25000, cuotas=12, total=300000))
    db.commit()

    r = client.get("/api/v1/bono/liquidacion/export?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "Resumen Renovaciones" in wb.sheetnames
    ws = wb["Resumen Renovaciones"]
    counts = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}
    assert counts[junior.name] == 2          # matcheadas
    assert counts["CCARO"] == 1              # sin matchear, por renovador_raw
    assert counts["TOTAL"] == 3              # solo julio
