"""Tests for the renovaciones (contract renewals) module.

Covers the derived fields (fecha_hasta = +1 year, total = cuota × 12), the RUT
normalization, the firm-lawyer selector, the monthly resumen, and delete perms.
"""
from datetime import date

import pytest

from app.core.security import create_access_token
from app.models.lawyer import Lawyer

ADMIN_RUT = "16021492-9"
LAWYER_RUT = "19643548-4"


@pytest.fixture
def admin(db):
    obj = Lawyer(rut=ADMIN_RUT, name="Carla Admin", role="admin")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@pytest.fixture
def abogado(db):
    obj = Lawyer(rut=LAWYER_RUT, name="Eduardo Venegas", role="lawyer", is_firm_lawyer=True)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def _h(rut):
    return {"Authorization": "Bearer " + create_access_token({"sub": rut})}


def test_abogados_selector_only_firm_active(client, admin, abogado, db):
    ext = Lawyer(rut="99999999-9", name="Contraparte Externa", role="lawyer", is_firm_lawyer=False)
    inact = Lawyer(rut="88888888-8", name="Sylvia Inactiva", role="lawyer", is_firm_lawyer=True, is_active=False)
    db.add_all([ext, inact])
    db.commit()
    r = client.get("/api/v1/renovaciones/abogados", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    names = {o["nombre"] for o in r.json()}
    assert "Eduardo Venegas" in names
    assert "Contraparte Externa" not in names  # not a firm lawyer
    assert "Sylvia Inactiva" not in names       # inactive


def test_abogados_selector_includes_procuradores(client, admin, abogado, db):
    """Procuradores (role='procurador', not is_firm_lawyer) are selectable for
    renovaciones even though they stay out of the firm-lawyer views."""
    proc = Lawyer(rut="20613995-1", name="CAMILA ANDREA CANALES COÑUENAO",
                  role="procurador", is_firm_lawyer=False, is_active=True)
    db.add(proc)
    db.commit()
    r = client.get("/api/v1/renovaciones/abogados", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    names = {o["nombre"] for o in r.json()}
    assert "CAMILA ANDREA CANALES COÑUENAO" in names
    assert "Eduardo Venegas" in names  # firm lawyers still included


def test_create_accepts_procurador(client, admin, db):
    proc = Lawyer(rut="19557032-9", name="CONSTANZA ANDREA CARO CORTEZ",
                  role="procurador", is_firm_lawyer=False, is_active=True)
    db.add(proc)
    db.commit()
    db.refresh(proc)
    r = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-PROC", "cliente_rut": "12345678-5",
        "cliente_nombre": "Cliente P", "lawyer_id": proc.id, "monto_cuota": 25000,
    })
    assert r.status_code == 201
    assert r.json()["lawyer_nombre"] == "CONSTANZA ANDREA CARO CORTEZ"


def test_create_derives_hasta_and_total(client, admin, abogado):
    r = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "1000012345",
        "cliente_rut": "17.098.014-k",
        "cliente_nombre": "Camila Cerda",
        "lawyer_id": abogado.id,
        "monto_cuota": 25000,
        "fecha_desde": "2026-07-21",
    })
    assert r.status_code == 201
    b = r.json()
    assert b["fecha_desde"] == "2026-07-21"
    assert b["fecha_hasta"] == "2027-07-21"   # +1 year
    assert b["total"] == 25000 * 12           # cuota × 12
    assert b["cuotas"] == 12
    assert b["cliente_rut"] == "17.098.014-K"  # normalized + formatted
    assert b["lawyer_nombre"] == "Eduardo Venegas"


def test_create_custom_amount(client, admin, abogado):
    r = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-9", "cliente_rut": "12345678-5",
        "cliente_nombre": "Otro Cliente", "lawyer_id": abogado.id, "monto_cuota": 40000,
    })
    assert r.status_code == 201
    assert r.json()["total"] == 40000 * 12


def test_create_defaults_amount_and_today(client, admin, abogado):
    r = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-1", "cliente_rut": "11111111-1",
        "cliente_nombre": "Cliente X", "lawyer_id": abogado.id,
    })
    assert r.status_code == 201
    b = r.json()
    assert b["monto_cuota"] == 25000            # default
    assert b["fecha_desde"] == date.today().isoformat()


def test_create_rejects_non_firm_lawyer(client, admin, db):
    ext = Lawyer(rut="77777777-7", name="Externo", role="lawyer", is_firm_lawyer=False)
    db.add(ext)
    db.commit()
    db.refresh(ext)
    r = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-2", "cliente_rut": "22222222-2",
        "cliente_nombre": "Y", "lawyer_id": ext.id,
    })
    assert r.status_code == 404


def test_create_rejects_duplicate_same_contract_same_period(client, admin, abogado):
    base = {"cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "numero_contrato": "C-A"}
    r1 = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "cliente_rut": "17.098.014-k", "fecha_desde": "2026-07-10"})
    assert r1.status_code == 201
    # SAME client + SAME contract in the SAME period → rejected (real duplicate).
    r2 = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "cliente_rut": "17098014-K", "fecha_desde": "2026-07-25"})
    assert r2.status_code == 409


def test_create_allows_same_rut_different_contract_same_period(client, admin, abogado):
    # A client can hold SEVERAL contracts, each a separate renewal in the same
    # period. Same RUT but a DIFFERENT contrato must be accepted.
    base = {"cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "cliente_rut": "17.098.014-k", "fecha_desde": "2026-07-10"}
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-A"}).status_code == 201
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-B"}).status_code == 201


def test_create_allows_same_rut_different_period(client, admin, abogado):
    # A client CAN renew again in a future period (e.g. next year).
    base = {"cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "cliente_rut": "17.098.014-k"}
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-2026", "fecha_desde": "2026-07-10"}).status_code == 201
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-2027", "fecha_desde": "2027-07-10"}).status_code == 201


def test_create_allows_new_client_rut(client, admin, abogado):
    base = {"cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "fecha_desde": "2026-07-10"}
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-A", "cliente_rut": "12345678-5"}).status_code == 201
    assert client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        **base, "numero_contrato": "C-B", "cliente_rut": "11111111-1"}).status_code == 201


def test_create_rejects_bad_amount_and_blank(client, admin, abogado):
    bad_amount = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-3", "cliente_rut": "3-3", "cliente_nombre": "Z",
        "lawyer_id": abogado.id, "monto_cuota": 0,
    })
    assert bad_amount.status_code == 400
    blank = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "  ", "cliente_rut": "3-3", "cliente_nombre": "Z",
        "lawyer_id": abogado.id,
    })
    assert blank.status_code == 422


def test_list_filter_and_resumen(client, admin, abogado):
    # Distinct client RUTs — a client can't repeat within a period.
    for i, m in enumerate(("2026-07-05", "2026-07-20", "2026-08-01")):
        client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
            "numero_contrato": f"C-{m}", "cliente_rut": f"1111111{i}-1",
            "cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "fecha_desde": m,
        })
    jul = client.get("/api/v1/renovaciones?periodo=2026-07", headers=_h(ADMIN_RUT)).json()["items"]
    assert len(jul) == 2
    res = client.get("/api/v1/renovaciones/resumen?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    assert res["count"] == 2
    assert res["total_cuotas"] == 50000
    assert res["total_anual"] == 50000 * 12


def test_list_paginates_total_pages_and_slices(client, admin, abogado):
    """The list is a bounded, paginated envelope: total/pages are correct and a
    second page returns the next (non-overlapping) slice."""
    for i, day in enumerate((5, 10, 15, 20, 25)):
        client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
            "numero_contrato": f"C-PAG-{i}", "cliente_rut": f"1000000{i}-1",
            "cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "fecha_desde": f"2026-09-{day:02d}",
        })

    p1 = client.get("/api/v1/renovaciones?periodo=2026-09&per_page=2&page=1", headers=_h(ADMIN_RUT)).json()
    assert p1["total"] == 5
    assert p1["pages"] == 3          # ceil(5 / 2)
    assert p1["page"] == 1 and p1["per_page"] == 2
    assert len(p1["items"]) == 2

    p2 = client.get("/api/v1/renovaciones?periodo=2026-09&per_page=2&page=2", headers=_h(ADMIN_RUT)).json()
    assert len(p2["items"]) == 2
    # No overlap between page 1 and page 2 (stable order → disjoint slices).
    assert {r["id"] for r in p1["items"]}.isdisjoint({r["id"] for r in p2["items"]})


def _make_xlsx(rows):
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AÑO2026"
    ws.append(["RUT", "NOMBRE", "N° CONTRATO", "CUOTAS", "DESDE", "HASTA", "RENOVADOR", "VALOR"])
    for r in rows:
        ws.append(r)
    tot = wb.create_sheet("TOTALES 2025")  # must be skipped
    tot.append(["ignore", "me"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_importar_excel_maps_and_dedups(client, admin, abogado):
    from datetime import datetime as dt
    rows = [
        ["17098014-k", "Cliente Uno", "C-100", 12, dt(2026, 1, 5), dt(2027, 1, 5), "EVENEGAS", 20000],
        ["12345678-5", "Cliente Dos", "C-101", 12, dt(2026, 2, 10), dt(2027, 2, 10), "MVERA", 25000],
        ["", "", "", "", "", "", "", ""],  # blank → skipped
    ]
    data = _make_xlsx(rows)
    r = client.post(
        "/api/v1/renovaciones/importar", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    )
    assert r.status_code == 200
    b = r.json()
    assert b["total_leidas"] == 2
    assert b["creadas"] == 2
    assert b["vinculadas"] == 1   # EVENEGAS → Eduardo Venegas
    assert b["como_texto"] == 1   # MVERA has no system lawyer

    # re-upload of the same file dedups everything
    r2 = client.post(
        "/api/v1/renovaciones/importar", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    )
    assert r2.json()["creadas"] == 0
    assert r2.json()["omitidas_duplicadas"] == 2

    # linked row shows the system lawyer; text-only row shows the raw renovador
    enero = client.get("/api/v1/renovaciones?periodo=2026-01", headers=_h(ADMIN_RUT)).json()["items"]
    assert enero[0]["renovador"] == "Eduardo Venegas"
    febrero = client.get("/api/v1/renovaciones?periodo=2026-02", headers=_h(ADMIN_RUT)).json()["items"]
    assert febrero[0]["renovador"] == "MVERA"
    assert febrero[0]["lawyer_id"] is None


def test_importar_matchea_username_de_procurador(client, admin, abogado, db):
    """El índice incluye procuradores: 'MVERA' → María José Vera Pichunante.
    Antes caía como texto crudo porque el índice solo tenía firm lawyers."""
    from datetime import datetime as dt
    db.add(Lawyer(rut="20568122-1", name="MARÍA JOSÉ VERA PICHUNANTE",
                  role="procurador", is_firm_lawyer=False, is_active=True))
    db.commit()
    rows = [
        ["12345678-5", "Cliente Dos", "C-200", 12, dt(2026, 2, 10), dt(2027, 2, 10), "MVERA", 25000],
    ]
    r = client.post(
        "/api/v1/renovaciones/importar", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", _make_xlsx(rows), _XLSX_MIME)},
    )
    assert r.status_code == 200
    assert r.json()["vinculadas"] == 1
    febrero = client.get("/api/v1/renovaciones?periodo=2026-02", headers=_h(ADMIN_RUT)).json()["items"]
    assert febrero[0]["renovador"] == "MARÍA JOSÉ VERA PICHUNANTE"
    assert febrero[0]["lawyer_id"] is not None


def test_listar_hojas_and_import_single_sheet(client, admin, abogado):
    from datetime import datetime as dt
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    a25 = wb.active
    a25.title = "AÑO2025"
    a25.append(["RUT", "NOMBRE", "N° CONTRATO", "CUOTAS", "DESDE", "HASTA", "RENOVADOR", "VALOR"])
    a25.append(["12345678-5", "Cliente 25", "C-25", 12, dt(2025, 3, 1), dt(2026, 3, 1), "EVENEGAS", 20000])
    a26 = wb.create_sheet("AÑO 2026")
    a26.append(["RUT", "NOMBRE", "N° CONTRATO", "CUOTAS", "DESDE", "HASTA", "RENOVADOR", "VALOR"])
    a26.append(["12345678-5", "Cliente 26a", "C-26a", 12, dt(2026, 3, 1), dt(2027, 3, 1), "EVENEGAS", 25000])
    a26.append(["12345678-5", "Cliente 26b", "C-26b", 12, dt(2026, 4, 1), dt(2027, 4, 1), "MVERA", 25000])
    tot = wb.create_sheet("TOTALES 2025")
    tot.append(["x", "y"])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    # the sheets endpoint lists importable sheets + row counts (skips TOTALES)
    hojas = client.post(
        "/api/v1/renovaciones/importar/hojas", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    ).json()
    by_name = {h["nombre"]: h["filas"] for h in hojas}
    assert by_name == {"AÑO2025": 1, "AÑO 2026": 2}
    assert "TOTALES 2025" not in by_name

    # import only AÑO 2026 → 2 rows, not the 2025 one
    r = client.post(
        "/api/v1/renovaciones/importar?hoja=AÑO%202026", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    ).json()
    assert r["creadas"] == 2
    assert client.get("/api/v1/renovaciones?periodo=2025-03", headers=_h(ADMIN_RUT)).json()["items"] == []


def test_import_unknown_sheet_404(client, admin, abogado):
    data = _make_xlsx([])
    r = client.post(
        "/api/v1/renovaciones/importar?hoja=NoExiste", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    )
    assert r.status_code == 404


def test_importar_trunca_campos_largos(client, admin, abogado):
    """A junk RUT cell (e.g. a name) must not crash the import — it's truncated
    to the column limit (Postgres enforces VARCHAR length; the real file has these)."""
    from datetime import datetime as dt
    long_rut = "EDISONDANILOGALLEGOSGRANDO-N"  # 28 chars, a name in the RUT column
    rows = [[long_rut, "Cliente Largo", "C-LARGO", 12, dt(2026, 5, 1), dt(2027, 5, 1), "EVENEGAS", 20000]]
    r = client.post(
        "/api/v1/renovaciones/importar", headers=_h(ADMIN_RUT),
        files={"archivo": ("reno.xlsx", _make_xlsx(rows), _XLSX_MIME)},
    )
    assert r.status_code == 200
    assert r.json()["creadas"] == 1
    row = client.get("/api/v1/renovaciones?periodo=2026-05", headers=_h(ADMIN_RUT)).json()["items"][0]
    assert len(row["cliente_rut"]) <= 20


def test_importar_requires_admin(client, abogado):
    data = _make_xlsx([])
    r = client.post(
        "/api/v1/renovaciones/importar", headers=_h(LAWYER_RUT),
        files={"archivo": ("reno.xlsx", data, _XLSX_MIME)},
    )
    assert r.status_code == 403


def test_recaudacion_by_year(client, admin, abogado):
    # Distinct client RUTs — two in January must be different clients.
    for i, fecha in enumerate(("2026-01-05", "2026-01-20", "2026-02-01")):
        client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
            "numero_contrato": f"R-{fecha}", "cliente_rut": f"2222222{i}-2",
            "cliente_nombre": "Cliente", "lawyer_id": abogado.id,
            "monto_cuota": 25000, "fecha_desde": fecha,
        })
    rec = client.get("/api/v1/renovaciones/recaudacion?anio=2026", headers=_h(ADMIN_RUT)).json()
    assert rec["anio"] == 2026
    assert len(rec["meses"]) == 12
    enero = rec["meses"][0]
    assert enero["mes"] == 1 and enero["count"] == 2
    assert enero["recaudacion"] == 50000
    assert enero["proyeccion_anual"] == 600000  # 50000 × 12
    assert rec["total_recaudacion"] == 75000
    assert rec["total_count"] == 3


def test_delete_permissions(client, admin, abogado, db):
    reno = client.post("/api/v1/renovaciones", headers=_h(ADMIN_RUT), json={
        "numero_contrato": "C-del", "cliente_rut": "12345678-5",
        "cliente_nombre": "Cliente", "lawyer_id": abogado.id,
    }).json()
    # another non-admin who didn't create it → 403
    other = Lawyer(rut="10101010-1", name="Otro", role="lawyer")
    db.add(other)
    db.commit()
    forbidden = client.delete(f"/api/v1/renovaciones/{reno['id']}", headers=_h("10101010-1"))
    assert forbidden.status_code == 403
    # admin → 204
    ok = client.delete(f"/api/v1/renovaciones/{reno['id']}", headers=_h(ADMIN_RUT))
    assert ok.status_code == 204


def test_resumen_incluye_por_abogado(client, db, admin):
    """El resumen mensual trae el desglose por abogado (cantidad + total anual),
    ordenado por total desc, e incluye renovaciones sin abogado como 'Sin asignar'."""
    from app.models.renovacion import Renovacion

    a = Lawyer(rut="70000000-1", name="Ana Abogada", role="lawyer", is_firm_lawyer=True)
    b = Lawyer(rut="70000000-2", name="Beto Abogado", role="lawyer", is_firm_lawyer=True)
    db.add_all([a, b])
    db.commit()
    db.refresh(a)
    db.refresh(b)

    def mk(lawyer_id, monto, day, renovador_raw=None):
        db.add(Renovacion(
            numero_contrato=f"C-{day}-{monto}", cliente_rut="1-9", cliente_nombre="Cliente",
            lawyer_id=lawyer_id, renovador_raw=renovador_raw,
            fecha_desde=date(2026, 7, day), fecha_hasta=date(2027, 7, day),
            monto_cuota=monto, cuotas=12, total=monto * 12,
        ))

    mk(a.id, 25000, 5)                       # Ana: 2 renovaciones → total 600.000
    mk(a.id, 25000, 10)
    mk(b.id, 40000, 12)                      # Beto: 1 → total 480.000
    mk(None, 10000, 15, renovador_raw="")    # sin abogado → "Sin asignar" 120.000
    db.commit()

    r = client.get("/api/v1/renovaciones/resumen?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 4
    pa = body["por_abogado"]
    # Ordenado por total desc: Ana (600k) > Beto (480k) > Sin asignar (120k)
    assert [x["lawyer_nombre"] for x in pa] == ["Ana Abogada", "Beto Abogado", "Sin asignar"]
    # comisión = cantidad × $10.400 (V2), NO el monto del contrato (total).
    assert pa[0]["cantidad"] == 2 and pa[0]["total"] == 600_000 and pa[0]["comision"] == 20_800
    assert pa[1]["cantidad"] == 1 and pa[1]["total"] == 480_000 and pa[1]["comision"] == 10_400
    assert pa[2]["lawyer_id"] is None and pa[2]["comision"] == 10_400
