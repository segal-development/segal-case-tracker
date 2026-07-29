"""Tests for the hitos (milestone → bonus) endpoints — Slice 1.

Covers the entry flow (mandatory PJUD evidence), admin approval (blocked without
evidence — the firm's "sin evidencia no se paga" rule), the per-lawyer resumen,
and the permission boundaries (a lawyer sees/approves nothing outside their own).
The hito-type catalog is normally seeded by migration 032; tests seed one row
directly since the test schema is built from models, not migrations.
"""
from datetime import date

import pytest

from app.core.security import create_access_token
from app.models.hito import Hito, HitoTipo, HITO_PENDIENTE
from app.models.lawyer import Lawyer

ADMIN_RUT = "16021492-9"
LAWYER_RUT = "19643548-4"


@pytest.fixture
def admin(db):
    obj = Lawyer(rut=ADMIN_RUT, name="Carla Admin", role="admin")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@pytest.fixture
def lawyer(db):
    obj = Lawyer(rut=LAWYER_RUT, name="Benjamín Lawyer", role="lawyer")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@pytest.fixture
def tipo(db):
    t = HitoTipo(
        code="pleno_prescripcion", label="Prescripción terminada", nivel="pleno",
        valor_bruto=8077, etapa_tramite="INGRESO EXCEPCIONES PRESCRIPCIÓN", orden=1,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


def _h(rut):
    return {"Authorization": "Bearer " + create_access_token({"sub": rut})}


def _create(client, headers, tipo_id, lawyer_id=None):
    data = {"hito_tipo_id": tipo_id, "fecha_hito": "2026-07-15", "rol_causa": "C-1-2026"}
    if lawyer_id is not None:
        data["lawyer_id"] = lawyer_id
    return client.post(
        "/api/v1/hitos", headers=headers, data=data,
        files={"evidencia": ("cap.png", b"\x89PNG_fake", "image/png")},
    )


def test_tipos_lists_catalog(client, db, admin, tipo):
    r = client.get("/api/v1/hitos/tipos", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    assert any(t["code"] == "pleno_prescripcion" and t["valor_bruto"] == 8077 for t in r.json())


def test_create_snapshots_value(client, db, admin, tipo):
    r = _create(client, _h(ADMIN_RUT), tipo.id)
    assert r.status_code == 201
    body = r.json()
    assert body["estado"] == HITO_PENDIENTE
    assert body["valor_bruto"] == 8077  # snapshot from the tipo
    assert body["tiene_evidencia"] is True


def test_create_without_evidence_ok(client, db, admin, tipo):
    # Evidence is optional now — creating without a file succeeds.
    r = client.post("/api/v1/hitos", headers=_h(ADMIN_RUT),
                    data={"hito_tipo_id": tipo.id, "fecha_hito": "2026-07-15"})
    assert r.status_code == 201
    assert r.json()["tiene_evidencia"] is False


def test_approve_without_evidence_ok(client, db, admin, lawyer, tipo):
    # A hito with no evidence can now be approved (evidence no longer mandatory).
    h = Hito(lawyer_id=lawyer.id, hito_tipo_id=tipo.id, valor_bruto=8077,
             fecha_hito=date(2026, 7, 15), estado=HITO_PENDIENTE)
    db.add(h)
    db.commit()
    db.refresh(h)
    r = client.post(f"/api/v1/hitos/{h.id}/aprobar", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    assert r.json()["estado"] == "aprobado"


def test_admin_approves_and_resumen_totals(client, db, admin, lawyer, tipo):
    hid = _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).json()["id"]
    r = client.post(f"/api/v1/hitos/{hid}/aprobar", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    assert r.json()["estado"] == "aprobado"

    r = client.get("/api/v1/hitos/resumen?periodo=2026-07", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    row = next(x for x in r.json() if x["lawyer_id"] == lawyer.id)
    assert row["aprobados"] == 1
    assert row["total_bruto"] == 8077


def test_regular_lawyer_cannot_approve(client, db, admin, lawyer, tipo):
    hid = _create(client, _h(LAWYER_RUT), tipo.id).json()["id"]
    r = client.post(f"/api/v1/hitos/{hid}/aprobar", headers=_h(LAWYER_RUT))
    assert r.status_code == 403  # require_admin


def test_lawyer_lists_only_own(client, db, admin, lawyer, tipo):
    _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=admin.id)   # admin's own
    _create(client, _h(LAWYER_RUT), tipo.id)                      # lawyer's own
    r = client.get("/api/v1/hitos", headers=_h(LAWYER_RUT))
    assert r.status_code == 200
    assert {x["lawyer_id"] for x in r.json()} == {lawyer.id}  # only their own


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _hitos_xlsx():
    import io
    import openpyxl
    from datetime import datetime as dt
    wb = openpyxl.Workbook()
    js = wb.active
    js.title = "HITOS JUNIOR"
    js.append(["HITOS H1 · ABOGADO JUNIOR"])                  # title row
    js.append(["Evidencia obligatoria: ..."])                # subtitle
    js.append([])                                            # blank
    js.append(["#", "Abogado AT", "Fecha registro", "PROCEDIMIENTO", "ROL causa",
               "Descripción", "ETAPA Sysgal", "TRAMITE Sysgal", "Aprobado", "Valor hito"])
    js.append([1, "Gonzalo Calderón", dt(2026, 7, 2), "RECURSO", "18.795.509-2", "DOC OK", "M1 BAJA", "Escrito", "SÍ", 2423])
    js.append([2, "Gonzalo Calderón", dt(2026, 7, 3), "EXHIBICION", "10.185.055-2", "OK", "M1 BAJA", "Escrito", "NO", 2423])
    p = wb.create_sheet("PARÁMETROS")  # must be skipped
    p.append(["no", "importar"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importar_hitos_hojas_y_mapeo(client, db, admin):
    jt = HitoTipo(code="junior_h1", label="Conversión preventiva → M1 Alta", nivel="junior", valor_bruto=2423, orden=1)
    lw = Lawyer(rut="19280895-2", name="Gonzalo Javier Calderón Maturana", role="lawyer", is_firm_lawyer=True)
    db.add_all([jt, lw])
    db.commit()
    data = _hitos_xlsx()

    # hojas: only the HITOS sheet (skips PARÁMETROS)
    hojas = client.post("/api/v1/hitos/importar/hojas", headers=_h(ADMIN_RUT),
                        files={"archivo": ("h.xlsx", data, _XLSX_MIME)}).json()
    assert [h["nombre"] for h in hojas] == ["HITOS JUNIOR"]
    assert hojas[0]["filas"] == 2

    # import → maps abogado by name, estado from Aprobado SÍ/NO
    r = client.post("/api/v1/hitos/importar", headers=_h(ADMIN_RUT),
                    files={"archivo": ("h.xlsx", data, _XLSX_MIME)}).json()
    assert r["creadas"] == 2
    assert r["aprobados"] == 1
    assert r["pendientes"] == 1

    # re-upload dedups
    r2 = client.post("/api/v1/hitos/importar", headers=_h(ADMIN_RUT),
                     files={"archivo": ("h.xlsx", data, _XLSX_MIME)}).json()
    assert r2["creadas"] == 0
    assert r2["omitidas_duplicadas"] == 2

    hitos = client.get("/api/v1/hitos?periodo=2026-07", headers=_h(ADMIN_RUT)).json()
    assert len(hitos) == 2
    assert {h["lawyer_id"] for h in hitos} == {lw.id}


def test_importar_hitos_requires_admin(client, db, lawyer):
    r = client.post("/api/v1/hitos/importar", headers=_h(LAWYER_RUT),
                    files={"archivo": ("h.xlsx", _hitos_xlsx(), _XLSX_MIME)})
    assert r.status_code == 403


def test_abogados_selector_lists_firm_lawyers(client, db, admin, lawyer):
    r = client.get("/api/v1/hitos/abogados", headers=_h(ADMIN_RUT))
    assert r.status_code == 200
    names = {a["nombre"] for a in r.json()}
    assert "Benjamín Lawyer" in names and "Carla Admin" in names


def test_admin_deletes_hito(client, db, admin, lawyer, tipo):
    hid = _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).json()["id"]
    forbidden = client.delete(f"/api/v1/hitos/{hid}", headers=_h(LAWYER_RUT))
    assert forbidden.status_code == 403           # non-admin cannot delete
    ok = client.delete(f"/api/v1/hitos/{hid}", headers=_h(ADMIN_RUT))
    assert ok.status_code == 204                  # admin can
    assert client.get("/api/v1/hitos?periodo=2026-07", headers=_h(ADMIN_RUT)).json() == []


def test_admin_edits_hito(client, db, admin, lawyer, tipo):
    hid = _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=admin.id).json()["id"]
    r = client.put(f"/api/v1/hitos/{hid}", headers=_h(ADMIN_RUT), json={
        "hito_tipo_id": tipo.id, "lawyer_id": lawyer.id,
        "fecha_hito": "2026-07-20", "rol_causa": "12.345.678-5", "descripcion": "editado",
    })
    assert r.status_code == 200
    b = r.json()
    assert b["lawyer_id"] == lawyer.id       # reassigned
    assert b["fecha_hito"] == "2026-07-20"
    assert b["descripcion"] == "editado"
    assert b["hito_tipo_id"] == tipo.id


def test_edit_hito_requires_admin(client, db, admin, lawyer, tipo):
    hid = _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).json()["id"]
    r = client.put(f"/api/v1/hitos/{hid}", headers=_h(LAWYER_RUT),
                   json={"hito_tipo_id": tipo.id, "fecha_hito": "2026-07-20"})
    assert r.status_code == 403


# --- No duplicate hito: (abogado, causa) strict ---


def test_create_rejects_duplicate_lawyer_causa(client, db, admin, lawyer, tipo):
    assert _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).status_code == 201
    # Same lawyer + same causa (C-1-2026) → rejected, regardless of tipo/fecha.
    assert _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).status_code == 409


def test_create_allows_same_lawyer_different_causa(client, db, admin, lawyer, tipo):
    assert _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id).status_code == 201
    r = client.post("/api/v1/hitos", headers=_h(ADMIN_RUT), data={
        "hito_tipo_id": tipo.id, "fecha_hito": "2026-07-15",
        "rol_causa": "C-999-2026", "lawyer_id": lawyer.id,
    }, files={"evidencia": ("cap.png", b"\x89PNG", "image/png")})
    assert r.status_code == 201


def test_create_allows_multiple_without_causa(client, db, admin, tipo):
    # Hitos without a causa can't collide → both allowed.
    base = {"hito_tipo_id": tipo.id, "fecha_hito": "2026-07-15"}
    assert client.post("/api/v1/hitos", headers=_h(ADMIN_RUT), data=base).status_code == 201
    assert client.post("/api/v1/hitos", headers=_h(ADMIN_RUT), data=base).status_code == 201


def test_edit_into_duplicate_causa_rejected(client, db, admin, lawyer, tipo):
    _create(client, _h(ADMIN_RUT), tipo.id, lawyer_id=lawyer.id)  # lawyer on C-1-2026
    other = client.post("/api/v1/hitos", headers=_h(ADMIN_RUT), data={
        "hito_tipo_id": tipo.id, "fecha_hito": "2026-07-15",
        "rol_causa": "C-2-2026", "lawyer_id": lawyer.id,
    }, files={"evidencia": ("cap.png", b"\x89PNG", "image/png")}).json()["id"]
    # Editing the second onto the first's causa collides.
    r = client.put(f"/api/v1/hitos/{other}", headers=_h(ADMIN_RUT), json={
        "hito_tipo_id": tipo.id, "lawyer_id": lawyer.id,
        "fecha_hito": "2026-07-15", "rol_causa": "C-1-2026",
    })
    assert r.status_code == 409


def test_importar_dedups_same_lawyer_causa_strict(client, db, admin):
    """Import cross-checks the strict (abogado, causa) key: same lawyer + same
    causa on different dates collapses to one row."""
    import io
    import openpyxl
    from datetime import datetime as dt
    jt = HitoTipo(code="junior_h1", label="X", nivel="junior", valor_bruto=2423, orden=1)
    lw = Lawyer(rut="19280895-2", name="Gonzalo Javier Calderón Maturana",
                role="lawyer", is_firm_lawyer=True)
    db.add_all([jt, lw])
    db.commit()
    wb = openpyxl.Workbook()
    js = wb.active
    js.title = "HITOS JUNIOR"
    js.append(["t"]); js.append(["s"]); js.append([])
    js.append(["#", "Abogado AT", "Fecha", "PROC", "ROL", "Desc", "ETAPA", "TRAM", "Aprob", "Valor"])
    js.append([1, "Gonzalo Calderón", dt(2026, 7, 2), "R", "18.795.509-2", "a", "M", "E", "NO", 2423])
    js.append([2, "Gonzalo Calderón", dt(2026, 7, 9), "R", "18.795.509-2", "b", "M", "E", "NO", 2423])
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post("/api/v1/hitos/importar", headers=_h(ADMIN_RUT),
                    files={"archivo": ("h.xlsx", buf.getvalue(), _XLSX_MIME)}).json()
    assert r["creadas"] == 1
    assert r["omitidas_duplicadas"] == 1
